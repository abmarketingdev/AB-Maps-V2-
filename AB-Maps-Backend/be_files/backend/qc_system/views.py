"""
Views for the QC System app.

Phase 2: Authentication views (qc_login, qc_logout, qc_me)
Phase 3: ContactViewSet, AgentViewSet
Phase 4: HistoryViewSet, SettingsViewSet
Phase 5: DashboardViewSet
Phase 6: ImportViewSet
Phase 9: Error handling, logging, query optimization
"""
import csv
import io
import json
import logging
from functools import wraps

from django.conf import settings as django_settings
from django.contrib.auth import authenticate
from django.core.exceptions import ValidationError as DjangoValidationError
from django.db import DatabaseError
from django.db import transaction
from django.db.models import Q, Count, Case, When, IntegerField, Exists, OuterRef
from django.http import HttpResponse
from django.utils import timezone
from django.utils.text import slugify
from datetime import datetime as dt_datetime

from rest_framework import status, permissions, viewsets
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken

from services.email_service import EmailService
from users.models import User
from campaigns.models import Campaign
from .models import (
    QCContact, QCHistory, QCSettings, ImportRecord, ContactAssignment,
    QCFavourite, QCCheckOff, QCTransferRequest, QCTransferRequestItem,
    SalesChiefNotifyLog,
)
from .permissions import IsQCUser, IsQCAdmin, IsQCEmployee
from . import qc_helpers
from .audit_helpers import record_audit
from .serializers import (
    QCLoginSerializer,
    QCLogoutSerializer,
    QCContactListSerializer,
    QCContactDetailSerializer,
    QCHistorySerializer,
    QCHistoryUpdateSerializer,
    QCApproveSerializer,
    BulkTransferSerializer,
    SalesChiefNotifySerializer,
    QCTransferRequestSerializer,
    QCTransferRequestDecisionSerializer,
    QCAgentSerializer,
    QCSettingsSerializer,
    ImportRecordSerializer,
    ImportRecordUpdateSerializer,
    ContactSaleDateSerializer,
    FileUploadSerializer,
)
from .gamification_services import award_events_for_qc_approve

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────
# Phase 9 — Configuration
# ──────────────────────────────────────────────

QC_IMPORT_MAX_ROWS = getattr(django_settings, 'QC_IMPORT_MAX_ROWS', 50_000)
QC_IMPORT_MAX_FILE_MB = getattr(django_settings, 'QC_IMPORT_MAX_FILE_MB', 10)


# ──────────────────────────────────────────────
# Phase 9 — Safe-view decorator (error handler)
# ──────────────────────────────────────────────

def qc_safe_view(func):
    """
    Decorator that wraps a DRF view/action in a try/except.
    Lets DRF/Django HTTP exceptions pass through (404, 403, etc.)
    and catches only truly unexpected errors.
    """
    from django.http import Http404
    from rest_framework.exceptions import APIException

    @wraps(func)
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except (Http404, APIException):
            # Let DRF handle these normally (404, 403, 401, validation, etc.)
            raise
        except DatabaseError:
            logger.exception("QC database error in %s", func.__name__)
            return Response(
                {'success': False, 'error': 'A database error occurred. Please try again.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        except Exception:
            logger.exception("QC unexpected error in %s", func.__name__)
            return Response(
                {'success': False, 'error': 'An unexpected error occurred.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
    return wrapper


# ──────────────────────────────────────────────
# Phase 2 — Authentication Views
# ──────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([permissions.AllowAny])
@qc_safe_view
def qc_login(request):
    """
    QC-specific login endpoint.

    Only allows users with employee_type='qc_emp' or admin_type='qc_admin'.
    Returns JWT tokens and user info formatted for the QC frontend.

    POST /api/qc/auth/login
    Body: { "username": "...", "password": "..." }
    """
    logger.info("QC login attempt for user: %s", request.data.get('username', '<missing>'))
    serializer = QCLoginSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(
            {'success': False, 'error': serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )

    username = serializer.validated_data['username']
    password = serializer.validated_data['password']

    user = authenticate(username=username, password=password)

    if not user:
        return Response(
            {'success': False, 'error': 'Invalid credentials.'},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    if not user.is_active:
        return Response(
            {'success': False, 'error': 'User account is disabled.'},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    # Gate: only QC users may log in here
    is_qc_employee = user.employee_type == 'qc_emp'
    is_qc_admin = user.admin_type == 'qc_admin' and user.is_superuser

    if not (is_qc_employee or is_qc_admin):
        return Response(
            {'success': False, 'error': 'Access denied. Only QC users can login here.'},
            status=status.HTTP_403_FORBIDDEN,
        )

    # Generate tokens
    refresh = RefreshToken.for_user(user)

    # Update last login
    user.last_login = timezone.now()
    user.save(update_fields=['last_login'])

    # Determine user type for frontend conditional rendering
    user_type = 'qc_admin' if is_qc_admin else 'qc_employee'

    record_audit(
        action_type='login',
        user=user,
        details={
            'user_type': user_type,
            'ip_address': request.META.get('REMOTE_ADDR', ''),
            'user_agent': request.META.get('HTTP_USER_AGENT', '')[:256],
            'login_time': timezone.now().isoformat(),
        },
    )

    logger.info("QC login SUCCESS: user=%s type=%s", user.username, user_type)

    return Response({
        'success': True,
        'data': {
            'access': str(refresh.access_token),
            'refresh': str(refresh),
            'expires_in': 28800,  # 8 hours
            'agent': {
                'id': str(user.id),
                'name': f"{user.first_name} {user.last_name}".strip() or user.username,
                'agentId': user.ab_person_id or str(user.id)[:8],
                'email': user.email,
                'userType': user_type,
            },
        },
    })


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def qc_logout(request):
    """
    QC logout — blacklists the refresh token.

    POST /api/qc/auth/logout
    Body: { "refresh": "<refresh_token>" }
    """
    serializer = QCLogoutSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(
            {'success': False, 'error': serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        token = RefreshToken(serializer.validated_data['refresh'])
        token.blacklist()
    except Exception:
        pass

    record_audit(
        action_type='logout',
        user=request.user if request.user.is_authenticated else None,
        details={'logout_time': timezone.now().isoformat()},
    )

    return Response({
        'success': True,
        'message': 'Successfully logged out.',
    })


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated, IsQCUser])
def qc_me(request):
    """
    Return the authenticated QC user's info.

    GET /api/qc/auth/me
    """
    user = request.user

    if user.admin_type == 'qc_admin' and user.is_superuser:
        user_type = 'qc_admin'
    else:
        user_type = 'qc_employee'

    return Response({
        'success': True,
        'data': {
            'id': str(user.id),
            'name': f"{user.first_name} {user.last_name}".strip() or user.username,
            'agentId': user.ab_person_id or str(user.id)[:8],
            'email': user.email,
            'userType': user_type,
        },
    })


# ──────────────────────────────────────────────
# Phase 3 — Contact ViewSet
# ──────────────────────────────────────────────

# Active (in-progress) statuses used throughout (includes tredje for existing data)
ACTIVE_STATUSES = [
    'til_behandling',
    'forste_oppring',
    'andre_oppring',
    'tredje_oppring',
]
# Columns to show in UI (3rd attempt column removed per QC/Blåkors)
ACTIVE_COLUMNS_UI = [
    'til_behandling',
    'forste_oppring',
    'andre_oppring',
]

# Multi-category filter: query param value -> QCContact boolean field
QC_CATEGORY_TO_FIELD = {
    'giverinspill': 'is_giverinspill',
    'si_opp': 'is_oppsigelse',
    'ris': 'is_ris',
    'noeytral': 'is_noeytral',
    'annen': 'is_annen',
    'positiv': 'is_positiv',
    'reservert': 'is_reservert',
    'utmeldt': 'is_utmeldt',
}


def _resolve_import_scope(request, user):
    """
    Resolve active import-list scope from:
    1) query params (`import_id` / `list_slug`)
    2) user settings (`selected_import_record`)
    3) fallback: no scoping
    """
    import_id = (request.query_params.get('import_id') or '').strip()
    if import_id:
        return {'import_record_id': import_id}

    list_slug = (request.query_params.get('list_slug') or '').strip().lower()
    if list_slug:
        return {'import_record__list_slug': list_slug}

    settings_obj = (
        QCSettings.objects
        .select_related('selected_import_record')
        .filter(user=user)
        .first()
    )
    if settings_obj and settings_obj.selected_import_record_id:
        return {'import_record_id': settings_obj.selected_import_record_id}

    return None


def _resolve_seller_names_from_ids(seller_id_strings):
    """
    Given a list of User UUID strings, return a set of name values that may be
    stored in QCContact.seller_name.  Covers the full name variants written at
    import time (employee.name / manager.name / first+last / username).
    Silently ignores UUIDs that don't match any user.
    """
    if not seller_id_strings:
        return set()

    valid_ids = []
    for raw in seller_id_strings:
        try:
            import uuid as _uuid
            valid_ids.append(_uuid.UUID(str(raw)))
        except (ValueError, AttributeError):
            pass

    if not valid_ids:
        return set()

    users = (
        User.objects
        .filter(id__in=valid_ids)
        .select_related('employee', 'manager')
    )

    names = set()
    for u in users:
        # Try linked profile name first (most likely to match CSV values)
        if u.employee_id:
            try:
                names.add(u.employee.name)
            except Exception:
                pass
        if u.manager_id:
            try:
                names.add(u.manager.name)
            except Exception:
                pass
        # Also include first+last and username as fallbacks
        full = f"{u.first_name} {u.last_name}".strip()
        if full:
            names.add(full)
        if u.username:
            names.add(u.username)

    names.discard('')
    return names


def _attach_check_off_scope_flags(contact, user):
    """
    Set _checked_off_default / _checked_off_siopp_ah on a contact for serializers
    when the queryset is not annotated (e.g. get_next, action responses).
    """
    scopes_present = set(
        QCCheckOff.objects.filter(user=user, contact_id=contact.pk).values_list(
            'scope', flat=True
        )
    )
    contact._checked_off_default = QCCheckOff.Scope.DEFAULT in scopes_present
    contact._checked_off_siopp_ah = QCCheckOff.Scope.SIOPP_AH in scopes_present


class ContactViewSet(viewsets.ModelViewSet):
    """
    ViewSet for QC contacts.

    list     GET    /api/qc/contacts/
    retrieve GET    /api/qc/contacts/:id/
    get_next GET    /api/qc/contacts/get_next/
    approve  PATCH  /api/qc/contacts/:id/approve/
    urgent   PATCH  /api/qc/contacts/:id/urgent/
    bulk_transfer PATCH /api/qc/contacts/bulk_transfer/
    """
    permission_classes = [permissions.IsAuthenticated, IsQCUser]
    lookup_field = 'pk'
    pagination_class = None  # Return all contacts (no page restriction)

    # ── Serializer selection ──────────────────
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return QCContactDetailSerializer
        return QCContactListSerializer

    # ── Queryset ──────────────────────────────
    def get_queryset(self):
        """
        Returns contacts visible to the current user.
        - QC employees: always only their own assigned contacts.
        - QC admins: all contacts by default; if show_mine=true, only their own.
        """
        user = self.request.user
        qs = QCContact.objects.select_related('assigned_to', 'campaign')

        is_qc_admin = user.admin_type == 'qc_admin' and user.is_superuser
        show_mine = self.request.query_params.get('show_mine', '').lower() in ('true', '1')

        if not is_qc_admin:
            qs = qs.filter(assigned_to=user)
        elif show_mine:
            qs = qs.filter(assigned_to=user)

        # Optional filters via query params
        status_filter = self.request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)

        campaign_filter = self.request.query_params.get('campaign')
        if campaign_filter:
            qs = qs.filter(campaign_id=campaign_filter)

        urgent_filter = self.request.query_params.get('urgent')
        if urgent_filter is not None and urgent_filter.lower() in ('true', '1'):
            qs = qs.filter(urgent=True)

        assigned_filter = self.request.query_params.get('assigned_to')
        if assigned_filter and is_qc_admin and not show_mine:
            qs = qs.filter(assigned_to_id=assigned_filter)

        # Import-list scope (query param override -> settings -> fallback no scope)
        import_scope_filter = _resolve_import_scope(self.request, user)
        if import_scope_filter:
            qs = qs.filter(**import_scope_filter)

        # Per-user favourite: annotate so serializer can show is_favourite without N+1
        qs = qs.annotate(
            _is_favourited_by_user=Exists(
                QCFavourite.objects.filter(user=user, contact=OuterRef('pk'))
            )
        )
        # Per-scope check-off: annotate for list/detail without N+1
        qs = qs.annotate(
            _checked_off_default=Exists(
                QCCheckOff.objects.filter(
                    user=user,
                    contact=OuterRef('pk'),
                    scope=QCCheckOff.Scope.DEFAULT,
                )
            ),
            _checked_off_siopp_ah=Exists(
                QCCheckOff.objects.filter(
                    user=user,
                    contact=OuterRef('pk'),
                    scope=QCCheckOff.Scope.SIOPP_AH,
                )
            ),
        )

        check_off_scope_raw = self.request.query_params.get('check_off_scope', '').strip().lower()
        if check_off_scope_raw and check_off_scope_raw not in QCCheckOff.Scope.values:
            raise ValidationError(
                {'check_off_scope': 'Invalid value. Use default or siopp_ah.'}
            )

        # Filter to only contacts the current user has favourited
        favourite_filter = self.request.query_params.get('favourite', '').lower()
        if favourite_filter in ('true', '1'):
            qs = qs.filter(_is_favourited_by_user=True)

        # Filter by check-off for a single scope (default if check_off_scope omitted)
        checked_off_param = self.request.query_params.get('checked_off', '').lower()
        filter_scope = (
            check_off_scope_raw if check_off_scope_raw in QCCheckOff.Scope.values
            else QCCheckOff.Scope.DEFAULT
        )
        if checked_off_param in ('true', '1'):
            if filter_scope == QCCheckOff.Scope.DEFAULT:
                qs = qs.filter(_checked_off_default=True)
            else:
                qs = qs.filter(_checked_off_siopp_ah=True)
        elif checked_off_param in ('false', '0'):
            if filter_scope == QCCheckOff.Scope.DEFAULT:
                qs = qs.filter(_checked_off_default=False)
            else:
                qs = qs.filter(_checked_off_siopp_ah=False)

        # Multi-category: show contacts that have this category (so same contact can appear in multiple lists)
        category_filter = self.request.query_params.get('category', '').strip().lower()
        if category_filter and category_filter in QC_CATEGORY_TO_FIELD:
            field_name = QC_CATEGORY_TO_FIELD[category_filter]
            qs = qs.filter(**{field_name: True})

        # Seller filter: ?seller_ids=<uuid>&seller_ids=<uuid> — resolves UUIDs to seller_name values
        seller_id_list = self.request.query_params.getlist('seller_ids')
        if seller_id_list:
            seller_names = _resolve_seller_names_from_ids(seller_id_list)
            if seller_names:
                qs = qs.filter(seller_name__in=seller_names)
            else:
                qs = qs.none()

        # Ordering: default oldest first (board). Use order=newest or sort=newest_first for reverse.
        order_param = self.request.query_params.get('order', '').lower()
        sort_param = self.request.query_params.get('sort', '').lower()
        if order_param in ('newest', '-created_at') or sort_param == 'newest_first':
            qs = qs.order_by('-created_at')
        else:
            qs = qs.order_by('user_added_import_date', 'created_at')

        return qs

    # ── Disable create / update / delete via standard REST ──
    def create(self, request, *args, **kwargs):
        return Response(
            {'success': False, 'error': 'Contacts are created via file import.'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    def update(self, request, *args, **kwargs):
        return Response(
            {'success': False, 'error': 'Use /approve or /urgent actions instead.'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    def partial_update(self, request, *args, **kwargs):
        return Response(
            {'success': False, 'error': 'Use /approve or /urgent actions instead.'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    def destroy(self, request, *args, **kwargs):
        return Response(
            {'success': False, 'error': 'Contacts cannot be deleted.'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    # ── list override for consistent response ──
    def list(self, request, *args, **kwargs):
        qs = self.filter_queryset(self.get_queryset())
        group_by = request.query_params.get('group_by_date', '').lower() in ('true', '1', 'yes')

        if group_by:
            contacts = list(qs)
            from collections import OrderedDict

            buckets = OrderedDict()
            for c in contacts:
                key = c.user_added_import_date.isoformat() if c.user_added_import_date else '_no_date'
                buckets.setdefault(key, []).append(c)
            sorted_keys = sorted((k for k in buckets if k != '_no_date'), key=lambda x: x)
            ordered = OrderedDict()
            for k in sorted_keys:
                ordered[k] = QCContactListSerializer(
                    buckets[k], many=True, context={'request': request}
                ).data
            if '_no_date' in buckets:
                ordered['_no_date'] = QCContactListSerializer(
                    buckets['_no_date'], many=True, context={'request': request}
                ).data
            return Response({
                'success': True,
                'count': len(contacts),
                'grouped_by_date': True,
                'date_keys': list(ordered.keys()),
                'data': dict(ordered),
            })

        expand_by_cat = request.query_params.get('expand_by_category', '').lower() in (
            'true', '1', 'yes',
        )
        if expand_by_cat:
            contacts = list(qs)
            category_param = request.query_params.get('category', '').strip().lower()
            slug_order = list(QC_CATEGORY_TO_FIELD.keys())

            # Serialize all contacts once with many=True, then expand in Python.
            # Avoids creating a new serializer instance per contact (was O(n) overhead).
            serialized = QCContactListSerializer(
                contacts, many=True, context={'request': request}
            ).data

            rows = []
            if category_param and category_param in QC_CATEGORY_TO_FIELD:
                for d in serialized:
                    rows.append({**d, 'list_category': category_param})
            else:
                for c, d in zip(contacts, serialized):
                    slugs = [
                        s for s in slug_order
                        if getattr(c, QC_CATEGORY_TO_FIELD[s], False)
                    ]
                    if slugs:
                        for s in slugs:
                            rows.append({**d, 'list_category': s})
                    else:
                        rows.append({**d, 'list_category': None})

            return Response({
                'success': True,
                'count': len(rows),
                'expanded_by_category': True,
                'data': rows,
            })

        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        contacts = list(qs)
        serializer = self.get_serializer(contacts, many=True)
        return Response({
            'success': True,
            'count': len(contacts),
            'data': serializer.data,
        })

    # ── retrieve override for consistent response ──
    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response({
            'success': True,
            'data': serializer.data,
        })

    # ── Custom actions ────────────────────────

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated, IsQCUser])
    @qc_safe_view
    def get_next(self, request):
        """
        Get the next contact for the current QC user (employee or admin).

        Optional query params:
        - prefer_status — valid ACTIVE_STATUSES value: try that column first (urgent then oldest).
        - only_prefer_status=true — with prefer_status: if that column is empty, return data null
          (no fallback to til_behandling / other columns).

        GET /api/qc/contacts/get_next/
        GET /api/qc/contacts/get_next/?prefer_status=forste_oppring&only_prefer_status=true

        Optional: check_off_scope=default|siopp_ah — exclude contacts checked off for that
        scope only (default: default).
        """
        user = request.user
        scope_raw = request.query_params.get('check_off_scope', '').strip().lower()
        if scope_raw and scope_raw not in QCCheckOff.Scope.values:
            return Response(
                {
                    'success': False,
                    'error': 'Invalid check_off_scope. Use default or siopp_ah.',
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        next_scope = scope_raw if scope_raw in QCCheckOff.Scope.values else QCCheckOff.Scope.DEFAULT

        base_qs = QCContact.objects.filter(
            assigned_to=user,
            status__in=ACTIVE_STATUSES,
        )
        import_scope_filter = _resolve_import_scope(request, user)
        if import_scope_filter:
            base_qs = base_qs.filter(**import_scope_filter)
        # Exclude contacts the current user has checked off for this scope
        base_qs = base_qs.exclude(
            pk__in=QCCheckOff.objects.filter(user=user, scope=next_scope).values_list(
                'contact_id', flat=True
            )
        )

        prefer_status = request.query_params.get('prefer_status', '').strip()
        only_prefer = request.query_params.get('only_prefer_status', '').lower() in (
            'true', '1', 'yes',
        )
        if prefer_status in ACTIVE_STATUSES:
            # Stay in same list: first try urgent in that status, then non-urgent (oldest first)
            urgent_in_preferred = base_qs.filter(
                status=prefer_status, urgent=True
            ).order_by('created_at').first()
            if urgent_in_preferred:
                _attach_check_off_scope_flags(urgent_in_preferred, user)
                return Response({
                    'success': True,
                    'data': QCContactListSerializer(
                        urgent_in_preferred, context={'request': request}
                    ).data,
                })
            contact_in_preferred = base_qs.filter(
                status=prefer_status, urgent=False
            ).order_by('created_at').first()
            if contact_in_preferred:
                _attach_check_off_scope_flags(contact_in_preferred, user)
                return Response({
                    'success': True,
                    'data': QCContactListSerializer(
                        contact_in_preferred, context={'request': request}
                    ).data,
                })
            if only_prefer:
                return Response({
                    'success': True,
                    'data': None,
                    'message': f'No contacts in {prefer_status}.',
                })
            # No contacts in preferred list — fall through to default priority

        # Priority 1: urgent contacts (oldest first)
        urgent_contact = base_qs.filter(urgent=True).order_by('created_at').first()
        if urgent_contact:
            _attach_check_off_scope_flags(urgent_contact, user)
            return Response({
                'success': True,
                'data': QCContactListSerializer(
                    urgent_contact, context={'request': request}
                ).data,
            })

        # Priority 2+: by status column (3rd attempt only for Blå Kors–type campaigns)
        third_campaign_ids = qc_helpers.campaign_ids_for_third_attempt()
        for col in ACTIVE_COLUMNS_UI:
            contact = (
                base_qs
                .filter(status=col, urgent=False)
                .order_by('created_at')
                .first()
            )
            if contact:
                _attach_check_off_scope_flags(contact, user)
                return Response({
                    'success': True,
                    'data': QCContactListSerializer(
                        contact, context={'request': request}
                    ).data,
                })
        if third_campaign_ids:
            contact = (
                base_qs
                .filter(
                    status='tredje_oppring',
                    urgent=False,
                    campaign_id__in=third_campaign_ids,
                )
                .order_by('created_at')
                .first()
            )
            if contact:
                _attach_check_off_scope_flags(contact, user)
                return Response({
                    'success': True,
                    'data': QCContactListSerializer(
                        contact, context={'request': request}
                    ).data,
                })

        return Response({
            'success': True,
            'data': None,
            'message': 'No contacts available.',
        })

    @action(detail=True, methods=['patch'], permission_classes=[permissions.IsAuthenticated, IsQCUser])
    @qc_safe_view
    def approve(self, request, pk=None):
        """
        Approve / complete a QC call.

        Business rules:
        - "Svarte" → svarteCategory REQUIRED → route to final column
        - "Ikke svar" / "Opptatt" → increment attemptCount → route to attempt column

        PATCH /api/qc/contacts/:id/approve/
        Body: { "qcResult": "...", "svarteCategory": "...", "siOpp": "...", "comment": "..." }
        """
        contact = self.get_object()
        user = request.user

        # QC employees may only approve contacts assigned to them; QC admins may approve any contact
        is_qc_admin = user.admin_type == 'qc_admin' and user.is_superuser
        if not is_qc_admin and contact.assigned_to != user:
            return Response(
                {'success': False, 'error': 'You can only approve contacts assigned to you.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = QCApproveSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {'success': False, 'error': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )

        data = serializer.validated_data
        qc_result = data['qcResult']
        svarte_category = data.get('svarteCategory')
        si_opp = data.get('siOpp')
        comment = data.get('comment', '')
        categories = data.get('categories') or []

        agent_name = f"{user.first_name} {user.last_name}".strip() or user.username

        # Update contact fields
        contact.qc_result = qc_result
        contact.svarte_category = svarte_category
        contact.si_opp = si_opp
        contact.comment = comment
        contact.qc_agent_name = agent_name
        contact.last_attempt_at = timezone.now()

        # Category flags — is_oppsigelse is additive (once set, never auto-cleared).
        # All other category flags are replaced on each approve so a correction
        # (e.g. Positiv → Nøytral) removes the old flag instead of leaving the
        # contact visible in both columns.
        _CATEGORY_FLAGS = [
            'is_giverinspill', 'is_ris', 'is_noeytral',
            'is_annen', 'is_positiv', 'is_reservert', 'is_utmeldt',
        ]
        if categories:
            # Explicit multi-select path: replace all category flags with the
            # current selection so stale flags from a previous approve are cleared.
            for _flag in _CATEGORY_FLAGS:
                setattr(contact, _flag, False)
            if 'si_opp' in categories:
                contact.is_oppsigelse = True
                contact.si_opp = contact.si_opp or 'JA'
            if 'giverinspill' in categories:
                contact.is_giverinspill = True
            if 'ris' in categories:
                contact.is_ris = True
            if 'noeytral' in categories:
                contact.is_noeytral = True
            if 'annen' in categories:
                contact.is_annen = True
            if 'positiv' in categories:
                contact.is_positiv = True
            if 'reservert' in categories:
                contact.is_reservert = True
            if 'utmeldt' in categories:
                contact.is_utmeldt = True
        else:
            if si_opp == 'JA':
                contact.is_oppsigelse = True
            if qc_result == 'Svarte' and svarte_category:
                # svarte_category values are mutually exclusive — clear all before
                # setting the new one so a correction doesn't leave ghost flags.
                for _flag in _CATEGORY_FLAGS:
                    setattr(contact, _flag, False)
                cat_to_flag = {
                    'giverinspill': 'is_giverinspill',
                    'negativ': 'is_ris',
                    'noeytral': 'is_noeytral',
                    'annen': 'is_annen',
                    'positiv': 'is_positiv',
                    'reservert': 'is_reservert',
                }
                if svarte_category in cat_to_flag:
                    setattr(contact, cat_to_flag[svarte_category], True)

        # Route status (primary column)
        if qc_result == 'Svarte':
            # Customer answered → final column
            if 'utmeldt' in categories:
                # Utmeldt overrides svarte_category routing — contact lands in utmeldt column only
                contact.status = 'utmeldt'
            else:
                status_map = {
                    'positiv': 'positiv_tilbakemelding',
                    'negativ': 'negativ_tilbakemelding',
                    'annen': 'other_inquiries',
                    'noeytral': 'noeytral_tilbakemelding',
                    'giverinspill': 'giverinspill',
                    'reservert': 'reservert',
                }
                contact.status = status_map.get(svarte_category, 'other_inquiries')
            contact.qc_approved_at = timezone.now()

            # Update assignment stats
            try:
                stats = ContactAssignment.objects.get(qc_employee=user)
                stats.active_assigned = max(stats.active_assigned - 1, 0)
                stats.completed_today += 1
                stats.save(update_fields=['active_assigned', 'completed_today', 'updated_at'])
            except ContactAssignment.DoesNotExist:
                pass
        else:
            # Customer didn't answer → increment attempt; 3rd column only for Blå Kors–type campaigns
            contact.attempt_count += 1
            if qc_helpers.campaign_allows_third_attempt(contact.campaign):
                amap = {1: 'forste_oppring', 2: 'andre_oppring', 3: 'tredje_oppring'}
                contact.status = amap.get(contact.attempt_count, 'tredje_oppring')
            else:
                amap = {1: 'forste_oppring', 2: 'andre_oppring'}
                contact.status = amap.get(contact.attempt_count, 'andre_oppring')

        contact.save()

        # ── Audit: call_outcome (always) ──────────────────────────────────────
        record_audit(
            action_type='call_outcome',
            user=user,
            contact=contact,
            details={
                'qc_result': qc_result,
                'svarte_category': svarte_category,
                'si_opp': si_opp,
                'is_utmeldt': bool(contact.is_utmeldt),
                'comment': comment[:500] if comment else '',
                'new_status': contact.status,
            },
        )
        # ── Audit: si_opp_flag (additive — only when JA) ─────────────────────
        if si_opp == 'JA' or 'si_opp' in categories:
            record_audit(
                action_type='si_opp_flag',
                user=user,
                contact=contact,
                details={
                    'qc_result': qc_result,
                    'svarte_category': svarte_category,
                    'comment': comment[:500] if comment else '',
                },
            )
        # ── Audit: utmeldt_flag ───────────────────────────────────────────────
        if 'utmeldt' in categories:
            record_audit(
                action_type='utmeldt_flag',
                user=user,
                contact=contact,
                details={
                    'qc_result': qc_result,
                    'svarte_category': svarte_category,
                    'comment': comment[:500] if comment else '',
                },
            )

        # Create immutable history entry (denormalize first/last name for NRC and display)
        history_entry = QCHistory.objects.create(
            contact=contact,
            customer_name=contact.customer_name,
            first_name=contact.first_name or '',
            last_name=contact.last_name or '',
            phone_number=contact.phone_number,
            qc_result=qc_result,
            svarte_category=svarte_category,
            si_opp=si_opp,
            comment=comment,
            qc_agent_name=agent_name,
            qc_agent=user,
            tid=timezone.now().strftime('%H:%M'),
        )

        # Phase 4: backend-driven gamification awards from approve outcomes.
        # Guarded in service by (user,event_type,contact,history_id) to avoid retry duplicates.
        try:
            award_events_for_qc_approve(
                user=user,
                contact=contact,
                history_entry=history_entry,
                qc_result=qc_result,
                svarte_category=svarte_category,
                si_opp=si_opp,
                comment=comment,
            )
        except Exception:
            # Gamification should never block core QC approve flow.
            logger.exception("Gamification award failed for contact=%s history=%s", contact.id, history_entry.id)

        logger.info(
            "Contact %s approved by %s – result=%s",
            contact.id, user.username, qc_result,
        )

        return Response({
            'success': True,
            'data': {
                'contact': QCContactListSerializer(contact).data,
                'historyEntry': QCHistorySerializer(history_entry).data,
            },
        })

    @action(detail=True, methods=['patch'], permission_classes=[permissions.IsAuthenticated, IsQCUser])
    def urgent(self, request, pk=None):
        """
        Toggle or set urgency on a contact.

        PATCH /api/qc/contacts/:id/urgent/
        Body: { "urgent": true, "urgent_message": "..." }
        """
        contact = self.get_object()

        urgent_val = request.data.get('urgent')
        if urgent_val is None:
            return Response(
                {'success': False, 'error': "'urgent' field is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        contact.urgent = bool(urgent_val)
        contact.urgent_message = request.data.get('urgent_message', '')
        contact.save()

        record_audit(
            action_type='urgent_set' if contact.urgent else 'urgent_cleared',
            user=request.user,
            contact=contact,
            details={
                'urgent': contact.urgent,
                'urgent_message': (contact.urgent_message or '')[:500],
            },
        )

        return Response({
            'success': True,
            'data': QCContactListSerializer(contact).data,
        })

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated, IsQCUser])
    @qc_safe_view
    def favourite(self, request, pk=None):
        """
        Toggle favourite for this contact for the current user (per-user favourites).
        POST /api/qc/contacts/:id/favourite/
        If the contact is not in the user's favourites, add it; otherwise remove it.
        Returns the updated contact and is_favourite flag.
        """
        contact = self.get_object()
        user = request.user

        fav, created = QCFavourite.objects.get_or_create(
            user=user,
            contact=contact,
            defaults={},
        )
        if created:
            is_favourite = True
        else:
            fav.delete()
            is_favourite = False

        # Re-annotate for serializer (single object may not have _is_favourited_by_user)
        contact._is_favourited_by_user = is_favourite
        return Response({
            'success': True,
            'data': {
                'contact': QCContactListSerializer(contact, context={'request': request}).data,
                'is_favourite': is_favourite,
            },
        })

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated, IsQCUser])
    @qc_safe_view
    def check_off(self, request, pk=None):
        """
        Toggle check-off for this contact for the current user (per-user, per scope).
        POST /api/qc/contacts/:id/check_off/
        Body (optional): { "scope": "default" | "siopp_ah" } — default is "default".

        If not checked off for that scope, add; otherwise remove.
        Returns contact, checked_off_by_me (legacy: default scope), and checked_off_by_scope.
        """
        contact = self.get_object()
        user = request.user

        data = request.data if isinstance(request.data, dict) else {}
        raw_scope = data.get('scope', QCCheckOff.Scope.DEFAULT)
        if raw_scope is None or raw_scope == '':
            scope = QCCheckOff.Scope.DEFAULT
        else:
            scope = str(raw_scope).strip().lower()
            if scope not in QCCheckOff.Scope.values:
                return Response(
                    {
                        'success': False,
                        'error': 'Invalid scope. Use "default" or "siopp_ah".',
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        check_off_obj, created = QCCheckOff.objects.get_or_create(
            user=user,
            contact=contact,
            scope=scope,
            defaults={},
        )
        if created:
            toggled_on = True
        else:
            check_off_obj.delete()
            toggled_on = False

        _attach_check_off_scope_flags(contact, user)

        return Response({
            'success': True,
            'data': {
                'contact': QCContactListSerializer(contact, context={'request': request}).data,
                'scope': scope,
                'checked_off_by_me': contact._checked_off_default,
                'checked_off_by_scope': {
                    'default': contact._checked_off_default,
                    'siopp_ah': contact._checked_off_siopp_ah,
                },
            },
        })

    @action(detail=True, methods=['patch'], permission_classes=[permissions.IsAuthenticated, IsQCAdmin])
    @qc_safe_view
    def update_sale_date(self, request, pk=None):
        """
        Admin-only: update the sale/batch date on a single contact.

        PATCH /api/qc/contacts/:id/update_sale_date/
        Body: { "userAddedImportDate": "2026-04-08" }   (YYYY-MM-DD, or null to clear)

        Only changes this one contact — other contacts in the same import are unaffected.
        """
        contact = self.get_object()
        serializer = ContactSaleDateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {'success': False, 'error': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )

        contact.user_added_import_date = serializer.validated_data['userAddedImportDate']
        contact.save(update_fields=['user_added_import_date'])

        logger.info(
            "Admin %s updated sale date on contact %s to %s",
            request.user.username, contact.id, contact.user_added_import_date,
        )

        return Response({
            'success': True,
            'data': QCContactListSerializer(contact, context={'request': request}).data,
        })

    @action(detail=False, methods=['patch'], permission_classes=[permissions.IsAuthenticated, IsQCUser])
    @qc_safe_view
    def bulk_transfer(self, request):
        """
        Bulk-transfer contacts to another QC agent.
        QC admin: immediate transfer.
        QC employee: creates a pending transfer request (no immediate reassignment).

        PATCH /api/qc/contacts/bulk_transfer/
        Body: { "contactIds": ["uuid1", ...], "targetAgentId": "uuid" }
        """
        serializer = BulkTransferSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {'success': False, 'error': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )

        contact_ids = serializer.validated_data['contactIds']
        target_agent_id = serializer.validated_data['targetAgentId']
        user = request.user
        is_qc_admin = user.admin_type == 'qc_admin' and user.is_superuser

        # Validate target agent exists and is a QC user (employee or admin)
        try:
            target_agent = User.objects.get(
                Q(id=target_agent_id),
                Q(employee_type='qc_emp') | Q(admin_type='qc_admin', is_superuser=True),
                is_active=True,
            )
        except User.DoesNotExist:
            return Response(
                {'success': False, 'error': 'Target agent not found or is not a QC user.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        contacts = QCContact.objects.filter(id__in=contact_ids)

        if not is_qc_admin:
            id_set = {str(i) for i in contact_ids}
            found_ids = {str(c.id) for c in contacts}
            if id_set != found_ids:
                return Response(
                    {
                        'success': False,
                        'error': 'One or more contact IDs were not found.',
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if contacts.exclude(assigned_to=user).exists():
                return Response(
                    {
                        'success': False,
                        'error': 'You can only transfer contacts assigned to you.',
                    },
                    status=status.HTTP_403_FORBIDDEN,
                )

            with transaction.atomic():
                transfer_request = QCTransferRequest.objects.create(
                    requested_by=user,
                    target_agent=target_agent,
                    status='pending',
                    requested_count=contacts.count(),
                )
                QCTransferRequestItem.objects.bulk_create([
                    QCTransferRequestItem(request=transfer_request, contact=c)
                    for c in contacts
                ])

            logger.info(
                "Transfer request created: %d contacts -> agent %s by %s (request=%s)",
                contacts.count(), target_agent.username, user.username, transfer_request.id,
            )
            return Response({
                'success': True,
                'data': {
                    'requestId': str(transfer_request.id),
                    'status': transfer_request.status,
                    'requestedCount': transfer_request.requested_count,
                    'targetAgent': {
                        'id': str(target_agent.id),
                        'name': f"{target_agent.first_name} {target_agent.last_name}".strip() or target_agent.username,
                    },
                },
            }, status=status.HTTP_201_CREATED)

        transferred_count = contacts.count()

        # Track previous agents for stats update
        previous_agents = set(
            contacts.exclude(assigned_to__isnull=True)
            .values_list('assigned_to_id', flat=True)
            .distinct()
        )

        contacts.update(assigned_to=target_agent)

        # Update assignment stats for target agent
        target_stats, _ = ContactAssignment.objects.get_or_create(qc_employee=target_agent)
        active_in_transfer = QCContact.objects.filter(
            id__in=contact_ids,
            status__in=ACTIVE_STATUSES,
        ).count()
        target_stats.total_assigned += transferred_count
        target_stats.active_assigned += active_in_transfer
        target_stats.last_assigned_at = timezone.now()
        target_stats.save()

        # Reduce active counts for previous agents
        for prev_id in previous_agents:
            if str(prev_id) == str(target_agent_id):
                continue
            try:
                prev_stats = ContactAssignment.objects.get(qc_employee_id=prev_id)
                prev_stats.active_assigned = max(prev_stats.active_assigned - active_in_transfer, 0)
                prev_stats.save(update_fields=['active_assigned', 'updated_at'])
            except ContactAssignment.DoesNotExist:
                pass

        logger.info(
            "Bulk transfer: %d contacts → agent %s by %s",
            transferred_count, target_agent.username, user.username,
        )

        record_audit(
            action_type='bulk_transfer',
            user=user,
            details={
                'count': transferred_count,
                'contact_ids': [str(cid) for cid in contact_ids][:50],
                'to_agent_id': str(target_agent.id),
                'to_agent_name': f"{target_agent.first_name} {target_agent.last_name}".strip() or target_agent.username,
                'to_agent_id_code': target_agent.ab_person_id or '',
            },
        )

        return Response({
            'success': True,
            'data': {
                'transferred': transferred_count,
                'targetAgent': {
                    'id': str(target_agent.id),
                    'name': f"{target_agent.first_name} {target_agent.last_name}".strip() or target_agent.username,
                },
            },
        })


# ──────────────────────────────────────────────
# Agent list (for bulk-transfer UI and dropdowns)
# ──────────────────────────────────────────────

class AgentViewSet(viewsets.ViewSet):
    """
    Read-only list of QC agents.

    GET /api/qc/agents/
    """
    permission_classes = [permissions.IsAuthenticated, IsQCUser]

    @qc_safe_view
    def list(self, request):
        agents = User.objects.filter(
            Q(employee_type='qc_emp') | Q(admin_type='qc_admin', is_superuser=True),
            is_active=True,
        ).select_related('employee').prefetch_related('assigned_qc_contacts')

        serializer = QCAgentSerializer(agents, many=True)
        return Response({
            'success': True,
            'count': agents.count(),
            'data': serializer.data,
        })


class TransferRequestViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Admin transfer-request queue and decision endpoints.

    list    GET  /api/qc/transfer-requests/
    detail  GET  /api/qc/transfer-requests/:id/
    accept  POST /api/qc/transfer-requests/:id/accept/
    decline POST /api/qc/transfer-requests/:id/decline/
    """
    permission_classes = [permissions.IsAuthenticated, IsQCAdmin]
    serializer_class = QCTransferRequestSerializer
    lookup_field = 'pk'

    def get_queryset(self):
        qs = QCTransferRequest.objects.select_related(
            'requested_by', 'target_agent', 'reviewed_by',
        ).prefetch_related('items__contact')

        status_filter = self.request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)

        requested_by_filter = self.request.query_params.get('requested_by')
        if requested_by_filter:
            qs = qs.filter(requested_by_id=requested_by_filter)

        target_agent_filter = self.request.query_params.get('target_agent')
        if target_agent_filter:
            qs = qs.filter(target_agent_id=target_agent_filter)

        date_from = self.request.query_params.get('date_from')
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)

        date_to = self.request.query_params.get('date_to')
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        return qs.order_by('-created_at')

    @qc_safe_view
    def list(self, request, *args, **kwargs):
        qs = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(qs, many=True)
        pending_count = QCTransferRequest.objects.filter(status='pending').count()
        return Response({
            'success': True,
            'count': qs.count(),
            'pendingCount': pending_count,
            'data': serializer.data,
        })

    @qc_safe_view
    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response({'success': True, 'data': serializer.data})

    @action(detail=True, methods=['post'])
    @qc_safe_view
    def accept(self, request, pk=None):
        tr = self.get_object()
        if tr.status != 'pending':
            return Response(
                {'success': False, 'error': 'Only pending requests can be accepted.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        contact_ids = list(tr.items.values_list('contact_id', flat=True))
        if not contact_ids:
            return Response(
                {'success': False, 'error': 'Transfer request has no contacts.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        contacts = QCContact.objects.filter(id__in=contact_ids)
        missing_count = len(contact_ids) - contacts.count()
        if missing_count > 0:
            return Response(
                {
                    'success': False,
                    'error': 'One or more contacts in this request no longer exist.',
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Strict consistency check: keep request pending if current assignee changed
        if tr.requested_by and contacts.exclude(assigned_to=tr.requested_by).exists():
            return Response(
                {
                    'success': False,
                    'error': 'Contact ownership changed after request creation. Request remains pending.',
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        target_agent = tr.target_agent
        target_agent_id = str(target_agent.id)
        transferred_count = contacts.count()
        previous_agents = set(
            contacts.exclude(assigned_to__isnull=True).values_list('assigned_to_id', flat=True).distinct()
        )

        with transaction.atomic():
            contacts.update(assigned_to=target_agent)

            target_stats, _ = ContactAssignment.objects.get_or_create(qc_employee=target_agent)
            active_in_transfer = QCContact.objects.filter(
                id__in=contacts.values_list('id', flat=True),
                status__in=ACTIVE_STATUSES,
            ).count()
            target_stats.total_assigned += transferred_count
            target_stats.active_assigned += active_in_transfer
            target_stats.last_assigned_at = timezone.now()
            target_stats.save()

            for prev_id in previous_agents:
                if str(prev_id) == target_agent_id:
                    continue
                try:
                    prev_stats = ContactAssignment.objects.get(qc_employee_id=prev_id)
                    prev_stats.active_assigned = max(prev_stats.active_assigned - active_in_transfer, 0)
                    prev_stats.save(update_fields=['active_assigned', 'updated_at'])
                except ContactAssignment.DoesNotExist:
                    pass

            tr.status = 'accepted'
            tr.reviewed_by = request.user
            tr.reviewed_at = timezone.now()
            tr.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'updated_at'])

        logger.info(
            "Transfer request accepted: request=%s requester=%s reviewer=%s target=%s transferred=%d",
            tr.id,
            tr.requested_by.username if tr.requested_by else '<deleted-user>',
            request.user.username,
            target_agent.username,
            transferred_count,
        )
        return Response({
            'success': True,
            'data': {
                'requestId': str(tr.id),
                'status': tr.status,
                'transferred': transferred_count,
            },
        })

    @action(detail=True, methods=['post'])
    @qc_safe_view
    def decline(self, request, pk=None):
        tr = self.get_object()
        if tr.status != 'pending':
            return Response(
                {'success': False, 'error': 'Only pending requests can be declined.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        decision_serializer = QCTransferRequestDecisionSerializer(data=request.data)
        if not decision_serializer.is_valid():
            return Response(
                {'success': False, 'error': decision_serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )

        tr.status = 'declined'
        tr.decline_reason = decision_serializer.validated_data.get('declineReason', '')
        tr.reviewed_by = request.user
        tr.reviewed_at = timezone.now()
        tr.save(update_fields=['status', 'decline_reason', 'reviewed_by', 'reviewed_at', 'updated_at'])

        logger.info(
            "Transfer request declined: request=%s requester=%s reviewer=%s target=%s requested_count=%d reason=%s",
            tr.id,
            tr.requested_by.username if tr.requested_by else '<deleted-user>',
            request.user.username,
            tr.target_agent.username,
            tr.requested_count,
            tr.decline_reason or '<empty>',
        )
        return Response({
            'success': True,
            'data': {
                'requestId': str(tr.id),
                'status': tr.status,
                'declineReason': tr.decline_reason,
            },
        })


# ──────────────────────────────────────────────
# Phase 4 — History ViewSet
# ──────────────────────────────────────────────

class HistoryViewSet(viewsets.ModelViewSet):
    """
    ViewSet for QC history entries.

    list     GET    /api/qc/history/
    retrieve GET    /api/qc/history/:id/
    update   PATCH  /api/qc/history/:id/        (comment only)
    export   GET    /api/qc/history/export/      (CSV download)
    """
    permission_classes = [permissions.IsAuthenticated, IsQCUser]
    serializer_class = QCHistorySerializer
    lookup_field = 'pk'

    def get_queryset(self):
        """
        - QC employees see only their own history entries.
        - QC admins see all history entries.
        Supports optional query-param filters: date, qc_result, agent.
        """
        user = self.request.user
        qs = QCHistory.objects.select_related('contact', 'qc_agent')

        is_admin = user.admin_type == 'qc_admin' and user.is_superuser
        if not is_admin:
            qs = qs.filter(qc_agent=user)

        # Filter by date
        date_filter = self.request.query_params.get('date')
        if date_filter:
            qs = qs.filter(date=date_filter)

        # Filter by date range
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)

        # Filter by QC result
        result_filter = self.request.query_params.get('qc_result')
        if result_filter:
            qs = qs.filter(qc_result=result_filter)

        # Filter by agent (admin only)
        agent_filter = self.request.query_params.get('agent')
        if agent_filter and is_admin:
            qs = qs.filter(qc_agent_id=agent_filter)

        # Filter by contact
        contact_filter = self.request.query_params.get('contact')
        if contact_filter:
            qs = qs.filter(contact_id=contact_filter)

        return qs.order_by('-created_at')

    # ── Disable create / delete ──
    def create(self, request, *args, **kwargs):
        return Response(
            {'success': False, 'error': 'History entries are created automatically via the approve action.'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    def destroy(self, request, *args, **kwargs):
        return Response(
            {'success': False, 'error': 'History entries cannot be deleted.'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    # ── list override ──
    def list(self, request, *args, **kwargs):
        qs = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(qs, many=True)
        return Response({
            'success': True,
            'count': qs.count(),
            'data': serializer.data,
        })

    # ── retrieve override ──
    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response({
            'success': True,
            'data': serializer.data,
        })

    # ── Update: only comment is editable ──
    def update(self, request, *args, **kwargs):
        return self._update_comment(request)

    def partial_update(self, request, *args, **kwargs):
        return self._update_comment(request)

    def _update_comment(self, request):
        instance = self.get_object()
        serializer = QCHistoryUpdateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {'success': False, 'error': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )

        old_comment = instance.comment
        instance.comment = serializer.validated_data['comment']
        instance.save(update_fields=['comment'])

        record_audit(
            action_type='comment_edit',
            user=request.user,
            contact=instance.contact if instance.contact_id else None,
            details={
                'history_entry_id': str(instance.id),
                'old_comment': (old_comment or '')[:500],
                'new_comment': (instance.comment or '')[:500],
            },
        )

        return Response({
            'success': True,
            'data': QCHistorySerializer(instance).data,
        })

    # ── Export as CSV ──
    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        Export history entries as a CSV file.
        Respects the same query-param filters as the list action.

        GET /api/qc/history/export/?date_from=YYYY-MM-DD&date_to=YYYY-MM-DD
        """
        qs = self.filter_queryset(self.get_queryset())

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = (
            f'attachment; filename="qc_history_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        )

        from .serializers import SVARTE_CATEGORY_TO_LABEL

        writer = csv.writer(response)
        writer.writerow([
            'ID', 'First Name', 'Last Name', 'Customer Name', 'Phone Number',
            'QC Result', 'Svarte Category', 'Si Opp',
            'Status Label', 'Oppsigelse', 'Display Labels',
            'Comment', 'QC Agent', 'Time', 'Date',
        ])

        for entry in qs.iterator():
            status_label = SVARTE_CATEGORY_TO_LABEL.get(entry.svarte_category or '', entry.svarte_category or '')
            is_oppsigelse = (entry.si_opp or '') == 'JA'
            display_labels = []
            if status_label:
                display_labels.append(status_label)
            if is_oppsigelse:
                display_labels.append('Oppsigelse')
            writer.writerow([
                str(entry.id),
                entry.first_name or '',
                entry.last_name or '',
                entry.customer_name,
                entry.phone_number,
                entry.qc_result,
                entry.svarte_category or '',
                entry.si_opp or '',
                status_label,
                'JA' if is_oppsigelse else 'NEI',
                '; '.join(display_labels) if display_labels else '',
                entry.comment,
                entry.qc_agent_name,
                entry.tid,
                str(entry.date),
            ])

        return response


# ──────────────────────────────────────────────
# Phase 4 — Settings ViewSet
# ──────────────────────────────────────────────

class SettingsViewSet(viewsets.ViewSet):
    """
    Per-user QC settings (retrieve & update).

    GET   /api/qc/settings/        → retrieve current user's settings
    PATCH /api/qc/settings/        → update current user's settings
    """
    permission_classes = [permissions.IsAuthenticated, IsQCUser]

    def list(self, request):
        """
        Retrieve the current user's QC settings.
        Auto-creates default settings if they don't exist yet.
        """
        settings_obj, _ = QCSettings.objects.select_related('selected_import_record').get_or_create(
            user=request.user
        )
        serializer = QCSettingsSerializer(settings_obj)
        return Response({
            'success': True,
            'data': serializer.data,
        })

    def create(self, request):
        """
        PATCH semantics mapped onto POST for convenience.
        Updates the current user's QC settings.
        """
        return self._update_settings(request)

    def _update_settings(self, request):
        settings_obj, _ = QCSettings.objects.select_related('selected_import_record').get_or_create(
            user=request.user
        )

        payload = request.data.copy()
        # Support both snake_case and camelCase API payloads.
        if payload.get('selectedImportId') and 'selected_import_record' not in payload:
            payload['selected_import_record'] = payload.get('selectedImportId')

        serializer = QCSettingsSerializer(settings_obj, data=payload, partial=True)
        if not serializer.is_valid():
            return Response(
                {'success': False, 'error': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )
        changed_fields = list(serializer.validated_data.keys())
        serializer.save()
        record_audit(
            action_type='settings_changed',
            user=request.user,
            details={'changed_fields': changed_fields},
        )
        return Response({
            'success': True,
            'data': serializer.data,
        })


# ──────────────────────────────────────────────
# Phase 5 — Dashboard ViewSet
# ──────────────────────────────────────────────

class DashboardViewSet(viewsets.ViewSet):
    """
    Dashboard statistics for QC users.

    GET /api/qc/dashboard/stats/
    """
    permission_classes = [permissions.IsAuthenticated, IsQCUser]

    @action(detail=False, methods=['get'])
    @qc_safe_view
    def stats(self, request):
        """
        Return campaign-specific dashboard statistics.

        Query params:
            campaign (required): Campaign UUID

        Response:
            - QC Admin: campaign_stats (all contacts in campaign) + personal_stats (own)
            - QC Employee: personal_stats (own) only
        """
        user = request.user
        today = timezone.now().date()
        is_admin = user.admin_type == 'qc_admin' and user.is_superuser

        # Validate campaign parameter
        campaign_id = request.query_params.get('campaign')
        if not campaign_id:
            return Response(
                {'success': False, 'error': 'campaign parameter is required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            campaign = Campaign.objects.get(pk=campaign_id)
        except Campaign.DoesNotExist:
            return Response(
                {'success': False, 'error': 'Campaign not found.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        all_statuses = [
            'til_behandling', 'forste_oppring', 'andre_oppring', 'tredje_oppring',
            'si_opp', 'negativ_tilbakemelding', 'positiv_tilbakemelding', 'other_inquiries',
            'noeytral_tilbakemelding', 'giverinspill', 'reservert',
        ]
        active_cols_campaign = qc_helpers.active_columns_for_campaign(campaign)

        response_data = {
            'campaign': {
                'id': str(campaign.id),
                'name': campaign.name,
            },
        }

        if is_admin:
            # Campaign-wide stats (all contacts in campaign)
            campaign_contacts_qs = QCContact.objects.filter(campaign=campaign)
            campaign_history_qs = QCHistory.objects.filter(
                contact__campaign=campaign,
                date=today,
            )

            campaign_agg = campaign_contacts_qs.aggregate(
                **{
                    s: Count(
                        Case(When(status=s, then=1), output_field=IntegerField())
                    )
                    for s in all_statuses
                }
            )
            campaign_column_counts = {s: campaign_agg[s] for s in all_statuses}

            campaign_history_agg = campaign_history_qs.aggregate(
                svarte=Count(Case(When(qc_result='Svarte', then=1), output_field=IntegerField())),
                ikke_svar=Count(Case(When(qc_result='Ikke svar', then=1), output_field=IntegerField())),
                opptatt=Count(Case(When(qc_result='Opptatt', then=1), output_field=IntegerField())),
                si_opp=Count(Case(When(si_opp='JA', then=1), output_field=IntegerField())),
            )

            campaign_completed_today = (
                campaign_history_agg['svarte']
                + campaign_history_agg['ikke_svar']
                + campaign_history_agg['opptatt']
            )

            campaign_total_active = sum(campaign_column_counts[s] for s in active_cols_campaign)
            campaign_total_contacts = campaign_contacts_qs.count()

            # Multi-category counts (one contact can be in several)
            campaign_category_agg = campaign_contacts_qs.aggregate(
                giverinspill=Count(Case(When(is_giverinspill=True, then=1), output_field=IntegerField())),
                si_opp=Count(Case(When(is_oppsigelse=True, then=1), output_field=IntegerField())),
                ris=Count(Case(When(is_ris=True, then=1), output_field=IntegerField())),
                noeytral=Count(Case(When(is_noeytral=True, then=1), output_field=IntegerField())),
                annen=Count(Case(When(is_annen=True, then=1), output_field=IntegerField())),
                positiv=Count(Case(When(is_positiv=True, then=1), output_field=IntegerField())),
                reservert=Count(Case(When(is_reservert=True, then=1), output_field=IntegerField())),
            )

            response_data['campaign_stats'] = {
                'completedToday': campaign_completed_today,
                'columns': campaign_column_counts,
                'active_columns': active_cols_campaign,
                'third_attempt_enabled': qc_helpers.campaign_allows_third_attempt(campaign),
                'categoryCounts': campaign_category_agg,
                'todayStats': {
                    'svarte': campaign_history_agg['svarte'],
                    'ikkeSvar': campaign_history_agg['ikke_svar'],
                    'opptatt': campaign_history_agg['opptatt'],
                    'siOpp': campaign_history_agg['si_opp'],
                },
                'totalActive': campaign_total_active,
                'totalContacts': campaign_total_contacts,
            }

        # Personal stats (user's own contacts in campaign)
        personal_contacts_qs = QCContact.objects.filter(
            campaign=campaign,
            assigned_to=user,
        )
        personal_history_qs = QCHistory.objects.filter(
            contact__campaign=campaign,
            qc_agent=user,
            date=today,
        )

        personal_agg = personal_contacts_qs.aggregate(
            **{
                s: Count(
                    Case(When(status=s, then=1), output_field=IntegerField())
                )
                for s in all_statuses
            }
        )
        personal_column_counts = {s: personal_agg[s] for s in all_statuses}

        personal_history_agg = personal_history_qs.aggregate(
            svarte=Count(Case(When(qc_result='Svarte', then=1), output_field=IntegerField())),
            ikke_svar=Count(Case(When(qc_result='Ikke svar', then=1), output_field=IntegerField())),
            opptatt=Count(Case(When(qc_result='Opptatt', then=1), output_field=IntegerField())),
            si_opp=Count(Case(When(si_opp='JA', then=1), output_field=IntegerField())),
        )

        personal_completed_today = (
            personal_history_agg['svarte']
            + personal_history_agg['ikke_svar']
            + personal_history_agg['opptatt']
        )

        personal_total_active = sum(personal_column_counts[s] for s in active_cols_campaign)

        # Multi-category counts for personal stats
        personal_category_agg = personal_contacts_qs.aggregate(
            giverinspill=Count(Case(When(is_giverinspill=True, then=1), output_field=IntegerField())),
            si_opp=Count(Case(When(is_oppsigelse=True, then=1), output_field=IntegerField())),
            ris=Count(Case(When(is_ris=True, then=1), output_field=IntegerField())),
            noeytral=Count(Case(When(is_noeytral=True, then=1), output_field=IntegerField())),
            annen=Count(Case(When(is_annen=True, then=1), output_field=IntegerField())),
            positiv=Count(Case(When(is_positiv=True, then=1), output_field=IntegerField())),
            reservert=Count(Case(When(is_reservert=True, then=1), output_field=IntegerField())),
        )

        try:
            settings_obj = QCSettings.objects.get(user=user)
            daily_goal = settings_obj.daily_goal
        except QCSettings.DoesNotExist:
            daily_goal = 100

        goal_progress = round((personal_completed_today / daily_goal) * 100, 1) if daily_goal > 0 else 0.0

        response_data['personal_stats'] = {
            'completedToday': personal_completed_today,
            'dailyGoal': daily_goal,
            'goalProgress': goal_progress,
            'columns': personal_column_counts,
            'active_columns': active_cols_campaign,
            'third_attempt_enabled': qc_helpers.campaign_allows_third_attempt(campaign),
            'categoryCounts': personal_category_agg,
            'todayStats': {
                'svarte': personal_history_agg['svarte'],
                'ikkeSvar': personal_history_agg['ikke_svar'],
                'opptatt': personal_history_agg['opptatt'],
                'siOpp': personal_history_agg['si_opp'],
            },
            'totalActive': personal_total_active,
        }

        return Response({'success': True, 'data': response_data})


# ──────────────────────────────────────────────
# Phase 6 — Import Management
# ──────────────────────────────────────────────

def _parse_csv(file_obj):
    """Parse a CSV file and return a list of dicts (rows)."""
    text = file_obj.read().decode('utf-8-sig')  # utf-8-sig handles BOM
    reader = csv.DictReader(io.StringIO(text))
    return [row for row in reader]


def _parse_xlsx(file_obj):
    """Parse an XLSX file and return a list of dicts (rows)."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(file_obj.read()), read_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(next(rows_iter))]
    except StopIteration:
        return []

    data = []
    for row in rows_iter:
        row_dict = {}
        for idx, val in enumerate(row):
            if idx < len(headers):
                row_dict[headers[idx]] = val if val is not None else ''
        data.append(row_dict)

    wb.close()
    return data


def _distribute_contacts(contacts, agent_ids):
    """
    Evenly distribute a list of QCContact instances among agents.
    Accepts both QC employees and QC admins.

    Returns a dict of {agent_id: count_assigned}.
    """
    if not agent_ids:
        return {}

    agents = list(
        User.objects.filter(
            Q(id__in=agent_ids),
            Q(employee_type='qc_emp') | Q(admin_type='qc_admin', is_superuser=True),
            is_active=True,
        )
    )

    if not agents:
        return {}

    result = {str(a.id): 0 for a in agents}

    # Round-robin assignment
    for idx, contact in enumerate(contacts):
        agent = agents[idx % len(agents)]
        contact.assigned_to = agent
        result[str(agent.id)] += 1

    # Bulk update in one query
    QCContact.objects.bulk_update(contacts, ['assigned_to'])

    # Update ContactAssignment stats for each agent
    for agent in agents:
        assigned_count = result[str(agent.id)]
        if assigned_count > 0:
            stats, _ = ContactAssignment.objects.get_or_create(qc_employee=agent)
            stats.total_assigned += assigned_count
            stats.active_assigned += assigned_count
            stats.last_assigned_at = timezone.now()
            stats.save(update_fields=['total_assigned', 'active_assigned', 'last_assigned_at', 'updated_at'])

    return result


class ImportViewSet(viewsets.ViewSet):
    """
    File import management for QC contacts.

    upload   POST   /api/qc/imports/upload/
    list     GET    /api/qc/imports/              (import history)
    retrieve GET    /api/qc/imports/:id/
    preview  POST   /api/qc/imports/preview/      (preview columns)
    """
    permission_classes = [permissions.IsAuthenticated, IsQCUser]

    def get_permissions(self):
        """
        Imports listing/detail should be available to all QC users
        so employees can choose list context. Mutating/admin actions
        remain admin-only.
        """
        admin_only_actions = {'upload', 'preview', 'update_meta', 'destroy'}
        if getattr(self, 'action', None) in admin_only_actions:
            permission_classes = [permissions.IsAuthenticated, IsQCAdmin]
        else:
            permission_classes = [permissions.IsAuthenticated, IsQCUser]
        return [permission() for permission in permission_classes]

    # ── Import History (list) ──
    def list(self, request):
        """
        List all import records, newest first.
        Supports query-param filters: campaign, date_from, date_to, status.

        GET /api/qc/imports/
        """
        qs = ImportRecord.objects.select_related('campaign', 'imported_by').all()

        campaign_filter = request.query_params.get('campaign')
        if campaign_filter:
            qs = qs.filter(campaign_id=campaign_filter)

        date_from = request.query_params.get('date_from')
        if date_from:
            qs = qs.filter(date__gte=date_from)

        date_to = request.query_params.get('date_to')
        if date_to:
            qs = qs.filter(date__lte=date_to)

        status_filter = request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)

        return Response({
            'success': True,
            'count': qs.count(),
            'data': ImportRecordSerializer(qs, many=True).data,
        })

    # ── Import Detail ──
    def retrieve(self, request, pk=None):
        """
        Retrieve a single import record.

        GET /api/qc/imports/:id/
        """
        try:
            record = ImportRecord.objects.select_related('campaign', 'imported_by').get(pk=pk)
        except ImportRecord.DoesNotExist:
            return Response(
                {'success': False, 'error': 'Import record not found.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        return Response({
            'success': True,
            'data': ImportRecordSerializer(record).data,
        })

    # ── Delete Import (admin only) ──
    def destroy(self, request, pk=None):
        """
        Admin-only: delete an import record and all contacts that belong to it.

        DELETE /api/qc/imports/:id/

        This is destructive and irreversible. All QCContact rows linked to this
        import are permanently deleted before the ImportRecord itself is removed.
        """
        try:
            record = ImportRecord.objects.get(pk=pk)
        except ImportRecord.DoesNotExist:
            return Response(
                {'success': False, 'error': 'Import record not found.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        with transaction.atomic():
            deleted_contacts, _ = QCContact.objects.filter(import_record=record).delete()
            record.delete()

        logger.info(
            "Admin %s deleted import %s ('%s') and %d contacts",
            request.user.username, pk, record.list_name, deleted_contacts,
        )

        return Response({
            'success': True,
            'data': {
                'deletedImportId': str(pk),
                'deletedContacts': deleted_contacts,
            },
        }, status=status.HTTP_200_OK)

    # ── Edit list name / sale date (admin only) ──
    @action(detail=True, methods=['patch'])
    @qc_safe_view
    def update_meta(self, request, pk=None):
        """
        Admin-only: edit the list name and/or sale date of an import.

        PATCH /api/qc/imports/:id/update_meta/
        Body (JSON, at least one field required):
            {
                "listName": "NF 08.04",                 (optional)
                "userAddedImportDate": "2026-04-08"     (optional, YYYY-MM-DD or null)
            }

        If userAddedImportDate is changed, the new date cascades to every QCContact
        in this import batch. listName change regenerates the list_slug.
        """
        try:
            record = ImportRecord.objects.select_related('campaign').get(pk=pk)
        except ImportRecord.DoesNotExist:
            return Response(
                {'success': False, 'error': 'Import record not found.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        serializer = ImportRecordUpdateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {'success': False, 'error': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )

        data = serializer.validated_data
        updated_fields = []

        with transaction.atomic():
            if 'listName' in data:
                new_name = data['listName']
                new_slug = slugify(new_name)

                # Enforce unique (campaign, list_slug) — skip the current record itself
                conflict_qs = ImportRecord.objects.filter(
                    campaign=record.campaign,
                    list_slug=new_slug,
                ).exclude(pk=record.pk)
                if conflict_qs.exists():
                    return Response(
                        {
                            'success': False,
                            'error': f"Another import in this campaign already uses the name '{new_name}' (slug: '{new_slug}').",
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                record.list_name = new_name
                record.list_slug = new_slug
                updated_fields.extend(['list_name', 'list_slug'])

            if 'userAddedImportDate' in data:
                new_date = data['userAddedImportDate']
                record.user_added_import_date = new_date
                updated_fields.append('user_added_import_date')

                # Cascade to all contacts in this import
                QCContact.objects.filter(import_record=record).update(
                    user_added_import_date=new_date,
                )

            record.save(update_fields=updated_fields)

        logger.info(
            "Admin %s updated import %s: fields=%s",
            request.user.username, pk, updated_fields,
        )

        return Response({
            'success': True,
            'data': ImportRecordSerializer(record).data,
        })

    # ── Preview file columns ──
    @action(detail=False, methods=['post'])
    def preview(self, request):
        """
        Preview the columns (headers) of an uploaded CSV/XLSX file
        so the frontend can build a column-mapping UI.

        POST /api/qc/imports/preview/
        Body (multipart): { "file": <file> }

        Response: { "columns": ["Navn", "Telefon", "Selger", ...], "sampleRows": [...] }
        """
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'success': False, 'error': 'File is required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        filename = file.name.lower()

        try:
            if filename.endswith('.csv'):
                rows = _parse_csv(file)
            elif filename.endswith('.xlsx'):
                rows = _parse_xlsx(file)
            else:
                return Response(
                    {'success': False, 'error': 'Invalid file format. Only CSV and XLSX are supported.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception as e:
            logger.exception("Error parsing preview file")
            return Response(
                {'success': False, 'error': f'Error parsing file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not rows:
            return Response(
                {'success': False, 'error': 'File is empty or has no data rows.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        columns = list(rows[0].keys())
        sample_rows = rows[:5]  # First 5 rows for preview

        return Response({
            'success': True,
            'data': {
                'columns': columns,
                'sampleRows': sample_rows,
                'totalRows': len(rows),
            },
        })

    # ── Upload & Import Contacts ──
    @action(detail=False, methods=['post'])
    @qc_safe_view
    def upload(self, request):
        """
        Import contacts from a CSV/XLSX file.

        POST /api/qc/imports/upload/
        Body (multipart/form-data):
            file          – CSV or XLSX file
            campaignId    – UUID of the target Campaign
            listName      – required import list name (e.g. list_nf)
            mappings      – JSON string: {"name": "Navn", "first_name": "Fornavn", "last_name": "Etternavn",
                            "phone": "Telefon", "seller": "Selger", "sales_date": "Salgsdato"}
            agentIds      – JSON string: ["uuid1", "uuid2", ...]  (optional)

        Column mapping defaults:
            name       → "Navn" (single full name; if used alone, split on first space into first/last)
            first_name → "Fornavn"
            last_name  → "Etternavn"
            phone      → "Telefon"
            seller     → "Selger"
            sales_date → (no default; omit key or map to "" to leave dates null)
                         Accepted date formats per row: YYYY-MM-DD, DD.MM.YYYY, DD/MM/YYYY
        """
        serializer = FileUploadSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {'success': False, 'error': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )

        file = serializer.validated_data['file']
        campaign_id = serializer.validated_data['campaignId']
        list_name = serializer.validated_data['listName'].strip()
        list_slug = slugify(list_name).replace('_', '-')
        if not list_slug:
            return Response(
                {'success': False, 'error': 'listName must contain letters or numbers.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Phase 9: file-size guard ──
        max_bytes = QC_IMPORT_MAX_FILE_MB * 1024 * 1024
        if file.size > max_bytes:
            logger.warning("Import rejected: file too large (%s bytes)", file.size)
            return Response(
                {'success': False, 'error': f'File exceeds the {QC_IMPORT_MAX_FILE_MB} MB limit.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Parse mappings JSON
        try:
            mappings = json.loads(serializer.validated_data['mappings'])
        except (json.JSONDecodeError, TypeError):
            mappings = {}

        # Parse agentIds JSON
        try:
            agent_ids = json.loads(serializer.validated_data['agentIds'])
        except (json.JSONDecodeError, TypeError):
            agent_ids = []

        # Validate campaign exists
        try:
            campaign = Campaign.objects.get(pk=campaign_id)
        except Campaign.DoesNotExist:
            return Response(
                {'success': False, 'error': 'Campaign not found.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        if ImportRecord.objects.filter(campaign=campaign, list_slug=list_slug).exists():
            return Response(
                {
                    'success': False,
                    'error': f"listName '{list_name}' already exists for this campaign.",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Create import record (status = processing)
        import_record = ImportRecord.objects.create(
            filename=file.name,
            list_name=list_name,
            list_slug=list_slug,
            campaign=campaign,
            count=0,
            status='Behandler',
            imported_by=request.user,
        )

        record_audit(
            action_type='import_started',
            user=request.user,
            details={
                'import_record_id': str(import_record.id),
                'filename': file.name,
                'list_name': list_name,
                'campaign_id': str(campaign.id),
                'campaign_name': campaign.name,
            },
        )

        # Parse file
        filename_lower = file.name.lower()
        try:
            if filename_lower.endswith('.csv'):
                rows = _parse_csv(file)
            elif filename_lower.endswith('.xlsx'):
                rows = _parse_xlsx(file)
            else:
                import_record.status = 'Feilet'
                import_record.save(update_fields=['status'])
                return Response(
                    {'success': False, 'error': 'Invalid file format. Only CSV and XLSX are supported.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception as e:
            logger.exception("Error parsing import file")
            import_record.status = 'Feilet'
            import_record.save(update_fields=['status'])
            return Response(
                {'success': False, 'error': f'Error parsing file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not rows:
            import_record.status = 'Feilet'
            import_record.save(update_fields=['status'])
            return Response(
                {'success': False, 'error': 'File is empty or has no data rows.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Phase 9: row-count guard
        if len(rows) > QC_IMPORT_MAX_ROWS:
            import_record.status = 'Feilet'
            import_record.save(update_fields=['status'])
            logger.warning(
                "Import rejected: too many rows (%d > %d)", len(rows), QC_IMPORT_MAX_ROWS,
            )
            return Response(
                {'success': False, 'error': f'File has {len(rows)} rows, maximum is {QC_IMPORT_MAX_ROWS}.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Column mapping with defaults (name / first_name+last_name for full name on cards and NRC)
        name_col = mappings.get('name', 'Navn')
        first_name_col = mappings.get('first_name', 'Fornavn')
        last_name_col = mappings.get('last_name', 'Etternavn')
        phone_col = mappings.get('phone', 'Telefon')
        seller_col = mappings.get('seller', 'Selger')
        sales_id_col = mappings.get('sales_id', '')
        contact_id_col = mappings.get('contact_id', '')
        sales_date_col = mappings.get('sales_date', '')

        _DATE_FMTS = ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y')

        def _parse_row_date(raw):
            raw = str(raw).strip()
            if not raw:
                return None
            for fmt in _DATE_FMTS:
                try:
                    return dt_datetime.strptime(raw, fmt).date()
                except ValueError:
                    continue
            return None

        # Create contacts
        contacts_created = []
        errors = []

        for row_idx, row in enumerate(rows, start=2):  # start=2 because row 1 is header
            # Prefer separate first/last columns; else use single "Navn" and split on first space
            first_name = str(row.get(first_name_col, '')).strip()
            last_name = str(row.get(last_name_col, '')).strip()
            single_name = str(row.get(name_col, '')).strip()
            if first_name or last_name:
                # Mapped first_name and/or last_name columns present
                pass
            elif single_name:
                # Only "Navn" column: split on first space into first_name, last_name
                parts = single_name.split(None, 1)
                first_name = parts[0]
                last_name = parts[1] if len(parts) > 1 else ''
            full_name = f"{first_name} {last_name}".strip()
            customer_name = full_name or single_name or 'Ukjent'

            phone_number = str(row.get(phone_col, '')).strip()
            seller_name = str(row.get(seller_col, '')).strip()
            sales_id = str(row.get(sales_id_col, '')).strip() if sales_id_col else ''
            contact_id_val = str(row.get(contact_id_col, '')).strip() if contact_id_col else ''
            contact_id_value = contact_id_val or None
            row_sales_date = _parse_row_date(row.get(sales_date_col, '')) if sales_date_col else None

            # Skip genuinely empty rows before fallback labels like "Ukjent"
            if not (first_name or last_name or single_name or phone_number):
                errors.append({
                    'row': row_idx,
                    'error': f"Missing both name and '{phone_col}'.",
                })
                continue

            contact = QCContact(
                customer_name=customer_name,
                first_name=first_name,
                last_name=last_name,
                phone_number=phone_number or '',
                seller_name=seller_name or '',
                sales_id=sales_id,
                contact_id=contact_id_value,
                campaign=campaign,
                import_record=import_record,
                status='til_behandling',
                user_added_import_date=row_sales_date,
            )
            contacts_created.append(contact)

        # Bulk create all contacts
        if contacts_created:
            QCContact.objects.bulk_create(contacts_created)

        # Distribute contacts among agents
        assignment_results = {}
        if agent_ids and contacts_created:
            # Re-fetch created contacts (bulk_create may not set pks for uuid)
            # since UUIDs are generated client-side via default, they are set
            assignment_results = _distribute_contacts(contacts_created, agent_ids)

        # Finalise import record
        import_record.count = len(contacts_created)
        import_record.status = 'Fullfort'
        import_record.save(update_fields=['count', 'status'])

        record_audit(
            action_type='import_completed',
            user=request.user,
            details={
                'import_record_id': str(import_record.id),
                'filename': file.name,
                'list_name': list_name,
                'campaign_id': str(campaign.id),
                'campaign_name': campaign.name,
                'contacts_created': len(contacts_created),
                'errors': len(errors),
            },
        )

        logger.info(
            "Import complete: file=%s contacts=%d errors=%d campaign=%s by=%s",
            file.name, len(contacts_created), len(errors),
            campaign.name, request.user.username,
        )

        return Response({
            'success': True,
            'data': {
                'importRecord': ImportRecordSerializer(import_record).data,
                'contactsCreated': len(contacts_created),
                'agentsAssigned': assignment_results,
                'errors': errors[:50],  # Cap error list at 50
                'totalErrors': len(errors),
            },
        }, status=status.HTTP_201_CREATED)


# ──────────────────────────────────────────────
# QC TODO Views
# ──────────────────────────────────────────────

def _qc_list_todos(request):
    """Shared list logic for QC todos."""
    from todos.models import Todo
    from todos.serializers import TodoSerializer

    user = request.user
    queryset = Todo.objects.filter(user=user).select_related(
        'assigned_by', 'related_address', 'related_campaign',
    )

    status_filter = request.query_params.get('status')
    if status_filter:
        queryset = queryset.filter(status=status_filter)

    priority_filter = request.query_params.get('priority')
    if priority_filter:
        queryset = queryset.filter(priority=priority_filter)

    is_assigned = request.query_params.get('is_admin_assigned')
    if is_assigned is not None:
        queryset = queryset.filter(is_admin_assigned=is_assigned.lower() == 'true')

    queryset = queryset.order_by('-priority', 'deadline', '-created_at')
    serializer = TodoSerializer(queryset, many=True)

    logger.info(
        "QC user %s retrieved %d todos (status=%s, priority=%s)",
        user.username, queryset.count(), status_filter, priority_filter,
    )

    return Response({
        'success': True,
        'count': queryset.count(),
        'results': serializer.data,
    })


def _qc_create_todo_personal(request):
    """Create a single personal todo for the calling user."""
    from todos.models import Todo
    from todos.serializers import TodoSerializer

    serializer = TodoSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(
            {'success': False, 'error': serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )

    serializer.save(
        user=request.user,
        assigned_by=None,
        is_admin_assigned=False,
    )

    logger.info("QC user %s created personal TODO: '%s'", request.user.username, serializer.data['title'])

    return Response(
        {'success': True, 'data': serializer.data},
        status=status.HTTP_201_CREATED,
    )


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, IsQCUser])
@qc_safe_view
def qc_get_todos(request):
    """
    List or create TODOs.

    GET  /api/qc/todos/
        Query params: status, priority, is_admin_assigned

    POST /api/qc/todos/
        QC employee: creates a personal todo (assigned to self).
        QC admin:
          - user_ids omitted / empty  → creates todo for themselves (personal).
          - user_ids present          → creates one todo per employee, marked
                                        is_admin_assigned=True.
        Body: { title, description?, priority?, deadline?, user_ids?: [...] }
    """
    if request.method == 'GET':
        return _qc_list_todos(request)

    # ── POST ──────────────────────────────────────────────────────────────
    from todos.models import Todo
    from todos.serializers import TodoSerializer
    from django.db import transaction as db_transaction

    user = request.user
    is_admin = user.admin_type == 'qc_admin' and user.is_superuser
    user_ids = request.data.get('user_ids', [])

    # ── Admin assigning to employees ──────────────────────────────────────
    if is_admin and user_ids:
        if not isinstance(user_ids, list):
            return Response(
                {'success': False, 'error': 'user_ids must be a list of UUIDs.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        title = request.data.get('title', '').strip()
        if not title:
            return Response(
                {'success': False, 'error': 'title is required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        employees = User.objects.filter(
            id__in=user_ids,
            employee_type='qc_emp',
            is_active=True,
        )

        found_ids = {str(e.id) for e in employees}
        invalid = [str(uid) for uid in user_ids if str(uid) not in found_ids]
        if invalid:
            return Response(
                {'success': False, 'error': f"Not valid QC employees: {', '.join(invalid)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        description = request.data.get('description', '')
        priority = request.data.get('priority', 'medium')
        if priority not in ('low', 'medium', 'high'):
            priority = 'medium'
        deadline = request.data.get('deadline')

        created = []
        with db_transaction.atomic():
            for emp in employees:
                todo = Todo.objects.create(
                    user=emp,
                    assigned_by=user,
                    is_admin_assigned=True,
                    title=title,
                    description=description,
                    priority=priority,
                    deadline=deadline,
                    status=Todo.Status.PENDING,
                )
                created.append({
                    'todo_id': str(todo.id),
                    'user_id': str(emp.id),
                    'username': emp.username,
                })

        logger.info(
            "QC admin %s assigned todo '%s' to %d employees via /todos/",
            user.username, title, len(created),
        )

        return Response({
            'success': True,
            'assigned_count': len(created),
            'created_todos': created,
        }, status=status.HTTP_201_CREATED)

    # ── Personal todo (employee always; admin when no user_ids) ───────────
    return _qc_create_todo_personal(request)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsQCUser])
@qc_safe_view
def qc_create_todo(request):
    """
    Create a personal TODO for the current QC user.
    (Backwards-compatible alias — POST /api/qc/todos/create/)
    """
    return _qc_create_todo_personal(request)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsQCUser])
@qc_safe_view
def qc_mark_todo_complete(request, todo_id):
    """
    Mark a specific TODO as completed for the current QC user.

    POST /api/qc/todos/<todo_id>/complete/
    """
    from todos.models import Todo
    from todos.serializers import TodoSerializer

    user = request.user

    try:
        todo = Todo.objects.select_related('assigned_by').get(id=todo_id, user=user)
    except Todo.DoesNotExist:
        logger.warning(
            "QC user %s tried to complete non-existent or unauthorized TODO %s",
            user.username, todo_id,
        )
        return Response(
            {'success': False, 'error': 'TODO not found or you do not have permission to complete it.'},
            status=status.HTTP_404_NOT_FOUND,
        )

    todo.status = Todo.Status.COMPLETED
    todo.completed_at = timezone.now()
    todo.save(update_fields=['status', 'completed_at', 'updated_at'])

    serializer = TodoSerializer(todo)
    logger.info("QC user %s marked TODO %s as complete: '%s'", user.username, todo.id, todo.title)

    return Response({'success': True, 'message': 'TODO marked as completed', 'data': serializer.data})


@api_view(['DELETE'])
@permission_classes([IsAuthenticated, IsQCUser])
@qc_safe_view
def qc_delete_todo(request, todo_id):
    """
    Delete a completed TODO for the current QC user.
    Only todos owned by the user and with status 'completed' can be deleted.

    DELETE /api/qc/todos/<todo_id>/
    """
    from todos.models import Todo

    user = request.user

    try:
        todo = Todo.objects.get(id=todo_id, user=user)
    except Todo.DoesNotExist:
        logger.warning(
            "QC user %s tried to delete non-existent or unauthorized TODO %s",
            user.username, todo_id,
        )
        return Response(
            {'success': False, 'error': 'TODO not found or you do not have permission to delete it.'},
            status=status.HTTP_404_NOT_FOUND,
        )

    if todo.status != Todo.Status.COMPLETED:
        return Response(
            {'success': False, 'error': 'Only completed todos can be deleted.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    title = todo.title
    todo.delete()
    logger.info("QC user %s deleted completed TODO %s: '%s'", user.username, todo_id, title)

    return Response({
        'success': True,
        'message': 'TODO deleted.',
    }, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsQCUser])
@qc_safe_view
def qc_bulk_complete_todos(request):
    """
    Bulk-complete TODOs for the current QC user.

    POST /api/qc/todos/bulk-complete/
    Body: { "todo_ids": ["uuid1", "uuid2", ...] }
    """
    from todos.models import Todo

    user = request.user
    todo_ids = request.data.get('todo_ids', [])

    if not todo_ids or not isinstance(todo_ids, list):
        return Response(
            {'success': False, 'error': 'todo_ids is required and must be a non-empty list.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    now = timezone.now()
    completable = Todo.objects.filter(
        id__in=todo_ids,
        user=user,
        status__in=[Todo.Status.PENDING, Todo.Status.IN_PROGRESS],
    )
    completed_ids = list(completable.values_list('id', flat=True))
    completable.update(status=Todo.Status.COMPLETED, completed_at=now)

    skipped_ids = [str(tid) for tid in todo_ids if tid not in completed_ids]

    logger.info(
        "QC user %s bulk-completed %d todos (%d skipped)",
        user.username, len(completed_ids), len(skipped_ids),
    )

    return Response({
        'success': True,
        'completed_count': len(completed_ids),
        'skipped_count': len(skipped_ids),
        'completed_ids': [str(i) for i in completed_ids],
        'skipped_ids': skipped_ids,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsQCAdmin])
@qc_safe_view
def qc_get_qc_employees(request):
    """
    List all QC employees (employee_type='qc_emp'). QC admin only.

    GET /api/qc/get-qc-employees/
    """
    employees = User.objects.filter(
        employee_type='qc_emp',
        is_active=True,
    ).select_related('employee').order_by('username')

    data = []
    for u in employees:
        data.append({
            'id': str(u.id),
            'username': u.username,
            'email': u.email,
            'first_name': u.first_name,
            'last_name': u.last_name,
            'ab_person_id': u.ab_person_id,
            'employee_id': str(u.employee.id) if u.employee else None,
            'employee_type': u.employee_type,
        })

    return Response({'success': True, 'count': len(data), 'results': data})


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsQCAdmin])
@qc_safe_view
def qc_assign_qc_employees(request):
    """
    QC admin assigns a task to QC employees.

    POST /api/qc/assign-qc-employees/
    Body: { "title", "description", "priority", "deadline", "user_ids": [...] }

    Only user_ids with employee_type='qc_emp' are accepted.
    Creates one Todo per employee with is_admin_assigned=True.
    """
    from django.db import transaction as db_transaction
    from todos.models import Todo

    title = request.data.get('title', '').strip()
    if not title:
        return Response(
            {'success': False, 'error': 'title is required.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    user_ids = request.data.get('user_ids', [])
    if not user_ids or not isinstance(user_ids, list):
        return Response(
            {'success': False, 'error': 'user_ids is required and must be a non-empty list.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    employees = User.objects.filter(
        id__in=user_ids,
        employee_type='qc_emp',
        is_active=True,
    ).select_related('employee')

    found_ids = {str(e.id) for e in employees}
    invalid = [str(uid) for uid in user_ids if str(uid) not in found_ids]
    if invalid:
        return Response(
            {'success': False, 'error': f"Not valid QC employees: {', '.join(invalid)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    description = request.data.get('description', '')
    priority = request.data.get('priority', 'medium')
    if priority not in ('low', 'medium', 'high'):
        priority = 'medium'
    deadline = request.data.get('deadline')

    created = []
    with db_transaction.atomic():
        for emp in employees:
            todo = Todo.objects.create(
                user=emp,
                assigned_by=request.user,
                is_admin_assigned=True,
                title=title,
                description=description,
                priority=priority,
                deadline=deadline,
                status=Todo.Status.PENDING,
            )
            created.append({
                'todo_id': str(todo.id),
                'user_id': str(emp.id),
                'username': emp.username,
            })

    logger.info(
        "QC admin %s assigned task '%s' to %d QC employees",
        request.user.username, title, len(created),
    )

    return Response(
        {
            'success': True,
            'message': f'Assigned task to {len(created)} QC employee(s)',
            'assigned_count': len(created),
            'created_todos': created,
        },
        status=status.HTTP_201_CREATED,
    )


# ──────────────────────────────────────────────
# Sales chiefs (QC admin): list + email digest
# ──────────────────────────────────────────────


def _qc_contact_digest_row(contact):
    """Flat dict for qc_sales_chief_contact_digest.html (English labels in template)."""
    from .serializers import _display_labels_contact, _status_label_from_contact

    first = (contact.first_name or '').strip()
    last = (contact.last_name or '').strip()
    full_name = f"{first} {last}".strip() or (contact.customer_name or '')
    assigned = ''
    if contact.assigned_to:
        u = contact.assigned_to
        assigned = f"{u.first_name} {u.last_name}".strip() or u.username
    labels = _display_labels_contact(contact)
    return {
        'id': str(contact.id),
        'contact_id': contact.contact_id or '',
        'full_name': full_name,
        'customer_name': contact.customer_name or '',
        'phone_number': contact.phone_number or '',
        'seller_name': contact.seller_name or '',
        'status': contact.status or '',
        'status_label': _status_label_from_contact(contact),
        'display_labels': ', '.join(labels) if labels else '',
        'qc_result': contact.qc_result or '',
        'svarte_category': contact.svarte_category or '',
        'si_opp': contact.si_opp or '',
        'attempt_count': str(contact.attempt_count),
        'urgent': contact.urgent,
        'urgent_message': (contact.urgent_message or '')[:2000],
        'comment': (contact.comment or '')[:8000],
        'qc_agent_name': contact.qc_agent_name or '',
        'campaign_name': contact.campaign.name if contact.campaign else '',
        'import_list_name': (
            contact.import_record.list_name if contact.import_record else ''
        ),
        'user_added_import_date': (
            str(contact.user_added_import_date) if contact.user_added_import_date else ''
        ),
        'assigned_qc_agent': assigned,
        'created_at': contact.created_at.isoformat() if contact.created_at else '',
        'last_attempt_at': (
            contact.last_attempt_at.isoformat() if contact.last_attempt_at else ''
        ),
        'qc_approved_at': (
            contact.qc_approved_at.isoformat() if contact.qc_approved_at else ''
        ),
    }


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsQCUser])
@qc_safe_view
def qc_sales_chiefs_list(request):
    """
    List active sales chiefs (User.is_sales_chief, active).

    GET /api/qc/sales-chiefs/
    Allowed for QC employees and QC admins (same as other QC read-only helpers).
    """
    chiefs = (
        User.objects.filter(is_sales_chief=True, is_active=True)
        .order_by('first_name', 'last_name', 'email')
    )
    data = []
    for u in chiefs:
        name = f"{u.first_name} {u.last_name}".strip() or u.username
        data.append({
            'id': str(u.id),
            'name': name,
            'email': u.email,
            'ab_person_id': u.ab_person_id,
        })
    return Response({
        'success': True,
        'count': len(data),
        'data': data,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsQCUser])
@qc_safe_view
def qc_sales_chief_team(request, chief_id):
    """
    List all team members under a specific sales chief.

    GET /api/qc/sales-chiefs/<chief_id>/team/
    Accessible by QC employees and QC admins.
    Returns managers and employees under the chief so the frontend
    can filter contacts by seller/agent.
    """
    from users.models import SalesChiefTeamMember

    try:
        chief = User.objects.get(id=chief_id, is_sales_chief=True, is_active=True)
    except User.DoesNotExist:
        return Response({'success': False, 'error': 'Sales chief not found.'}, status=404)

    memberships = (
        SalesChiefTeamMember.objects
        .filter(sales_chief=chief)
        .select_related('member')
        .order_by('role', 'member__first_name', 'member__last_name')
    )

    team = []
    for m in memberships:
        u = m.member
        team.append({
            'user_id': str(u.id),
            'name': f"{u.first_name} {u.last_name}".strip() or u.username,
            'email': u.email,
            'username': u.username,
            'ab_person_id': u.ab_person_id,
            'role': m.role,
        })

    return Response({
        'success': True,
        'chief': {
            'id': str(chief.id),
            'name': f"{chief.first_name} {chief.last_name}".strip() or chief.username,
            'email': chief.email,
        },
        'count': len(team),
        'team': team,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsQCUser])
@qc_safe_view
def qc_sales_chiefs_notify(request):
    """
    Email a digest of QC contacts to one sales chief.

    POST /api/qc/sales-chiefs/notify/
    Body: { "salesChiefId": "<uuid>", "contactIds": ["<uuid>", ...] }

    Allowed for QC employees and QC admins. Audit log stores the authenticated user as sent_by.
    """
    serializer = SalesChiefNotifySerializer(data=request.data)
    if not serializer.is_valid():
        return Response(
            {'success': False, 'error': serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )

    chief_id = serializer.validated_data['salesChiefId']
    contact_ids = serializer.validated_data['contactIds']

    try:
        chief = User.objects.get(
            id=chief_id,
            is_sales_chief=True,
            is_active=True,
        )
    except User.DoesNotExist:
        return Response(
            {
                'success': False,
                'error': 'Sales chief not found or is not an active sales chief.',
            },
            status=status.HTTP_404_NOT_FOUND,
        )

    if not (chief.email or '').strip():
        return Response(
            {'success': False, 'error': 'Sales chief has no email address.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    contacts = (
        QCContact.objects.filter(id__in=contact_ids)
        .select_related('campaign', 'import_record', 'assigned_to')
    )
    by_id = {str(c.id): c for c in contacts}
    missing = [str(cid) for cid in contact_ids if str(cid) not in by_id]
    if missing:
        return Response(
            {
                'success': False,
                'error': 'One or more contact IDs were not found.',
                'missing_ids': missing[:50],
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    def _seller_sort_key(contact):
        """Sort numerically when seller_name is a plain integer ID (e.g. 1500, 2067),
        fall back to case-insensitive alphabetical for named sellers."""
        name = (contact.seller_name or '').strip()
        try:
            return (0, int(name), '')
        except ValueError:
            return (1, 0, name.lower())

    ordered = sorted([by_id[str(cid)] for cid in contact_ids], key=_seller_sort_key)
    rows = [_qc_contact_digest_row(c) for c in ordered]

    chief_name = f"{chief.first_name} {chief.last_name}".strip() or chief.username
    sent_by = (
        f"{request.user.first_name} {request.user.last_name}".strip()
        or request.user.username
    )
    if request.user.email:
        sent_by = f"{sent_by} <{request.user.email}>"

    ok = EmailService.send_qc_sales_chief_digest(
        recipient_email=chief.email.strip(),
        chief_display_name=chief_name,
        contact_rows=rows,
        sent_by_display=sent_by,
    )
    if not ok:
        return Response(
            {
                'success': False,
                'error': 'Failed to send email. Check server logs and SMTP configuration.',
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    # Save audit log
    sender_name = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username
    SalesChiefNotifyLog.objects.create(
        sent_by=request.user,
        sent_by_name=sender_name,
        sales_chief=chief,
        sales_chief_name=chief_name,
        sales_chief_email=chief.email.strip(),
        contact_count=len(rows),
        contacts_snapshot=rows,
    )

    logger.info(
        "QC sales chief digest: %d contacts emailed to chief %s by %s",
        len(rows),
        chief.username,
        request.user.username,
    )

    return Response({
        'success': True,
        'emailed_count': len(rows),
        'chief': {
            'id': str(chief.id),
            'email': chief.email,
            'name': chief_name,
        },
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsQCAdmin])
@qc_safe_view
def qc_sales_chiefs_notify_log(request):
    """
    Audit log of sales chief email digests (who emailed which chief).

    GET /api/qc/sales-chiefs/notify-log/

    Sender filter (auth User.id only — same table as JWT user; optional Employee.id):
        (none)              – only rows where sent_by = the authenticated user
        sent_by_id=<uuid>   – only rows sent by that user (User.id or Employee.id FK)
        include_all_senders=true – all senders (same as legacy global list)

    Other query params:
        sales_chief_id  – filter by recipient sales chief (User UUID)
        date_from       – YYYY-MM-DD (sent_at >=)
        date_to         – YYYY-MM-DD (sent_at <=)

    Returns each log entry with the full contacts_snapshot so the admin
    can see exactly which cards were included in each email.
    """
    qs = SalesChiefNotifyLog.objects.select_related('sent_by', 'sales_chief').all()

    chief_filter = request.query_params.get('sales_chief_id')
    if chief_filter:
        qs = qs.filter(sales_chief_id=chief_filter)

    sent_by_param = (request.query_params.get('sent_by_id') or '').strip()
    include_all = request.query_params.get('include_all_senders', '').lower() in (
        'true', '1', 'yes',
    )
    if sent_by_param:
        try:
            sender = User.objects.get(pk=sent_by_param)
        except (User.DoesNotExist, DjangoValidationError):
            try:
                sender = User.objects.get(employee_id=sent_by_param)
            except (User.DoesNotExist, DjangoValidationError):
                return Response(
                    {
                        'success': False,
                        'error': 'sent_by_id must be a valid auth User id or Employee id.',
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
        qs = qs.filter(sent_by_id=sender.id)
    elif not include_all:
        qs = qs.filter(sent_by=request.user)

    date_from = request.query_params.get('date_from')
    if date_from:
        qs = qs.filter(sent_at__date__gte=date_from)

    date_to = request.query_params.get('date_to')
    if date_to:
        qs = qs.filter(sent_at__date__lte=date_to)

    data = []
    for log in qs:
        data.append({
            'id': str(log.id),
            'sent_at': log.sent_at.isoformat(),
            'sent_at_display': log.sent_at.strftime('%d.%m.%Y %H:%M'),
            'sent_by': {
                'id': str(log.sent_by_id) if log.sent_by_id else None,
                'name': log.sent_by_name,
            },
            'sales_chief': {
                'id': str(log.sales_chief_id) if log.sales_chief_id else None,
                'name': log.sales_chief_name,
                'email': log.sales_chief_email,
            },
            'contact_count': log.contact_count,
            'contacts': [
                {
                    'id': c.get('id'),
                    'full_name': c.get('full_name') or c.get('customer_name'),
                    'seller_name': c.get('seller_name'),
                    'sale_date': c.get('user_added_import_date'),
                    'status_label': c.get('status_label'),
                    'svarte_category': c.get('svarte_category'),
                    'qc_result': c.get('qc_result'),
                    'si_opp': c.get('si_opp'),
                    'campaign_name': c.get('campaign_name'),
                    'import_list_name': c.get('import_list_name'),
                    'comment': c.get('comment'),
                    'qc_agent_name': c.get('qc_agent_name'),
                }
                for c in (log.contacts_snapshot or [])
            ],
        })

    return Response({
        'success': True,
        'count': len(data),
        'data': data,
    })


# ──────────────────────────────────────────────
# Admin Agent Board — view any employee's board
# ──────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsQCAdmin])
@qc_safe_view
def qc_admin_agent_board(request):
    """
    Admin-only: view a QC employee's board exactly as the employee sees it.

    Returns the same data and response shape as GET /api/qc/contacts/, but
    scoped to a specific agent. Favourites and check-off annotations are
    resolved from the *target agent's* perspective, not the admin's.
    The import-scope settings fallback also uses the target agent's saved list.

    GET /api/qc/admin/agent-board/

    Required:
        agent_id        — UUID of the QC employee: auth User.id, or Employee.id
                          (same value as get-qc-employees `id` or `employee_id`)

    All ContactViewSet list filters work identically:
        status, campaign, urgent, import_id, list_slug,
        favourite, checked_off, check_off_scope, category,
        order, sort, group_by_date, expand_by_category

    Pagination: unlike GET /api/qc/contacts/, this endpoint does not paginate;
    `page_size` is ignored. `count` is the number of matching contacts (or
    expanded rows when expand_by_category=true).

    When `campaign` is set, `campaign_metrics` summarizes that campaign for
    this agent (totals are not limited by import_scope / favourites / etc.,
    except `contacts_matching_filters` which reflects the full queryset `qs`).
    """
    from collections import OrderedDict

    # ── 1. Resolve target agent ──────────────────────────────────────────────
    agent_id = (request.query_params.get('agent_id') or '').strip()
    if not agent_id:
        return Response(
            {'success': False, 'error': 'agent_id is required.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        target_agent = User.objects.get(
            id=agent_id,
            employee_type='qc_emp',
            is_active=True,
        )
    except User.DoesNotExist:
        try:
            target_agent = User.objects.get(
                employee_id=agent_id,
                employee_type='qc_emp',
                is_active=True,
            )
        except User.DoesNotExist:
            return Response(
                {'success': False, 'error': 'Agent not found or is not an active QC employee.'},
                status=status.HTTP_404_NOT_FOUND,
            )

    # ── 2. Base queryset: only this agent's contacts ─────────────────────────
    qs = QCContact.objects.select_related('assigned_to', 'campaign').filter(
        assigned_to=target_agent,
    )

    # ── 3. Apply all standard filters ────────────────────────────────────────
    status_filter = request.query_params.get('status')
    if status_filter:
        qs = qs.filter(status=status_filter)

    campaign_filter = request.query_params.get('campaign')
    if campaign_filter:
        qs = qs.filter(campaign_id=campaign_filter)

    urgent_filter = request.query_params.get('urgent', '').lower()
    if urgent_filter in ('true', '1'):
        qs = qs.filter(urgent=True)

    # Import-list scope: explicit params override; fallback to *target agent's* settings
    import_scope_filter = _resolve_import_scope(request, target_agent)
    if import_scope_filter:
        qs = qs.filter(**import_scope_filter)

    # Multi-category filter
    category_filter = request.query_params.get('category', '').strip().lower()
    if category_filter and category_filter in QC_CATEGORY_TO_FIELD:
        qs = qs.filter(**{QC_CATEGORY_TO_FIELD[category_filter]: True})

    # Seller filter: ?seller_ids=<uuid>&seller_ids=<uuid>
    seller_id_list = request.query_params.getlist('seller_ids')
    if seller_id_list:
        seller_names = _resolve_seller_names_from_ids(seller_id_list)
        if seller_names:
            qs = qs.filter(seller_name__in=seller_names)
        else:
            qs = qs.none()

    # ── 4. Annotate favourites & check-offs from target agent's POV ──────────
    qs = qs.annotate(
        _is_favourited_by_user=Exists(
            QCFavourite.objects.filter(user=target_agent, contact=OuterRef('pk'))
        ),
        _checked_off_default=Exists(
            QCCheckOff.objects.filter(
                user=target_agent,
                contact=OuterRef('pk'),
                scope=QCCheckOff.Scope.DEFAULT,
            )
        ),
        _checked_off_siopp_ah=Exists(
            QCCheckOff.objects.filter(
                user=target_agent,
                contact=OuterRef('pk'),
                scope=QCCheckOff.Scope.SIOPP_AH,
            )
        ),
    )

    # Validate check_off_scope
    check_off_scope_raw = request.query_params.get('check_off_scope', '').strip().lower()
    if check_off_scope_raw and check_off_scope_raw not in QCCheckOff.Scope.values:
        return Response(
            {'success': False, 'error': 'Invalid check_off_scope. Use default or siopp_ah.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Favourite filter (show only contacts the target agent has favourited)
    favourite_filter = request.query_params.get('favourite', '').lower()
    if favourite_filter in ('true', '1'):
        qs = qs.filter(_is_favourited_by_user=True)

    # Checked-off filter (reflect target agent's check-offs)
    checked_off_param = request.query_params.get('checked_off', '').lower()
    filter_scope = (
        check_off_scope_raw if check_off_scope_raw in QCCheckOff.Scope.values
        else QCCheckOff.Scope.DEFAULT
    )
    if checked_off_param in ('true', '1'):
        if filter_scope == QCCheckOff.Scope.DEFAULT:
            qs = qs.filter(_checked_off_default=True)
        else:
            qs = qs.filter(_checked_off_siopp_ah=True)
    elif checked_off_param in ('false', '0'):
        if filter_scope == QCCheckOff.Scope.DEFAULT:
            qs = qs.filter(_checked_off_default=False)
        else:
            qs = qs.filter(_checked_off_siopp_ah=False)

    # ── 5. Ordering (identical to ContactViewSet) ─────────────────────────────
    order_param = request.query_params.get('order', '').lower()
    sort_param = request.query_params.get('sort', '').lower()
    if order_param in ('newest', '-created_at') or sort_param == 'newest_first':
        qs = qs.order_by('-created_at')
    else:
        qs = qs.order_by('user_added_import_date', 'created_at')

    # ── 5b. Per-campaign metrics (when campaign filter is present) ───────────
    campaign_metrics = None
    if campaign_filter:
        try:
            camp = Campaign.objects.get(pk=campaign_filter)
            camp_name = camp.name
        except Campaign.DoesNotExist:
            camp_name = None
        base_campaign_qs = QCContact.objects.filter(
            assigned_to=target_agent,
            campaign_id=campaign_filter,
        )
        campaign_metrics = {
            'campaign_id': str(campaign_filter),
            'campaign_name': camp_name,
            # All contacts for this agent in this campaign (ignores import list, status, etc.)
            'contacts_total_in_campaign': base_campaign_qs.count(),
            # Pipeline / "active" work still in call rounds (same notion as board columns)
            'active_pipeline_in_campaign': base_campaign_qs.filter(
                status__in=ACTIVE_STATUSES,
            ).count(),
            # Rows matching the same filters as `data` (before category expansion)
            'contacts_matching_filters': qs.count(),
        }

    # ── 6. Agent metadata block ───────────────────────────────────────────────
    agent_display_name = (
        f"{target_agent.first_name} {target_agent.last_name}".strip()
        or target_agent.username
    )
    today = timezone.now().date()
    completed_today = QCHistory.objects.filter(qc_agent=target_agent, date=today).count()

    try:
        assignment_stats = ContactAssignment.objects.get(qc_employee=target_agent)
        active_assigned = assignment_stats.active_assigned
        total_assigned = assignment_stats.total_assigned
    except ContactAssignment.DoesNotExist:
        active_assigned = 0
        total_assigned = 0

    try:
        agent_settings = QCSettings.objects.get(user=target_agent)
        daily_goal = agent_settings.daily_goal
    except QCSettings.DoesNotExist:
        daily_goal = 100

    agent_meta = {
        'id': str(target_agent.id),
        'name': agent_display_name,
        'agentId': target_agent.ab_person_id or str(target_agent.id)[:8],
        'email': target_agent.email,
        'daily_goal': daily_goal,
        'completed_today': completed_today,
        'active_assigned': active_assigned,
        'total_assigned': total_assigned,
    }

    # ── 7. Render response (same modes as ContactViewSet.list) ────────────────
    group_by = request.query_params.get('group_by_date', '').lower() in ('true', '1', 'yes')
    if group_by:
        contacts = list(qs)
        buckets = OrderedDict()
        for c in contacts:
            key = c.user_added_import_date.isoformat() if c.user_added_import_date else '_no_date'
            buckets.setdefault(key, []).append(c)
        sorted_keys = sorted(k for k in buckets if k != '_no_date')
        ordered = OrderedDict()
        for k in sorted_keys:
            ordered[k] = QCContactListSerializer(
                buckets[k], many=True, context={'request': request}
            ).data
        if '_no_date' in buckets:
            ordered['_no_date'] = QCContactListSerializer(
                buckets['_no_date'], many=True, context={'request': request}
            ).data
        return Response({
            'success': True,
            'agent': agent_meta,
            'campaign_metrics': campaign_metrics,
            'count': len(contacts),
            'grouped_by_date': True,
            'date_keys': list(ordered.keys()),
            'data': dict(ordered),
        })

    expand_by_cat = request.query_params.get('expand_by_category', '').lower() in (
        'true', '1', 'yes',
    )
    if expand_by_cat:
        contacts = list(qs)
        slug_order = list(QC_CATEGORY_TO_FIELD.keys())
        rows = []

        def _row(contact, list_cat):
            return QCContactListSerializer(
                contact,
                context={'request': request, 'list_category': list_cat},
            ).data

        for c in contacts:
            if category_filter and category_filter in QC_CATEGORY_TO_FIELD:
                rows.append(_row(c, category_filter))
                continue
            slugs = [s for s in slug_order if getattr(c, QC_CATEGORY_TO_FIELD[s], False)]
            if slugs:
                for s in slugs:
                    rows.append(_row(c, s))
            else:
                rows.append(_row(c, None))

        return Response({
            'success': True,
            'agent': agent_meta,
            'campaign_metrics': campaign_metrics,
            'count': len(rows),
            'expanded_by_category': True,
            'data': rows,
        })

    # Flat list
    contacts = list(qs)
    return Response({
        'success': True,
        'agent': agent_meta,
        'campaign_metrics': campaign_metrics,
        'count': len(contacts),
        'data': QCContactListSerializer(contacts, many=True, context={'request': request}).data,
    })
